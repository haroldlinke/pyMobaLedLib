# -*- coding: utf-8 -*-
#
#         Workbook
#
# * Version: 4.03
# * Author: Harold Linke
# * Date: January 8, 2022
# * Copyright: Harold Linke 2021
# *
# *
# * MobaLedCheckColors on Github: https://github.com/haroldlinke/MobaLedCheckColors
# *
# *  
# * https://github.com/Hardi-St/MobaLedLib
# *
# * MobaLedCheckColors is free software: you can redistribute it and/or modify
# * it under the terms of the GNU General Public License as published by
# * the Free Software Foundation, either version 3 of the License, or
# * (at your option) any later version.
# *
# * MobaLedCheckColors is distributed in the hope that it will be useful,
# * but WITHOUT ANY WARRANTY; without even the implied warranty of
# * MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# * GNU General Public License for more details.
# *
# * You should have received a copy of the GNU General Public License
# * along with this program.  if not, see <http://www.gnu.org/licenses/>.
# *
# *
# ***************************************************************************

#------------------------------------------------------------------------------
# CHANGELOG:
# 2021-12-23 v4.01 HL: - Inital Version converted by VB2PY based on MLL V3.1.0
# 2022-01-07 v4.02 HL: - Else:, ByRef check done - first PoC release
# 2022-01-08 V4.03 HL: - added Workbook Save and Load
# 2022-04-02 LX4.18 HL: update to MLL V3.1.0F4, removed keyword "exit_on_error" and annotation for global variables
# 2022-04-03 LX4.19 HL: added platformcheck

#from collections import namedtuple
import logging
import locale
import platform
import pathlib
import re
import json
import copy
import openpyxl
from tksheet import Sheet #, num2alpha
import tkinter as tk
from tkinter import ttk

from vb2py.vbconstants import *
from vb2py.vbfunctions import *
from vb2py import vbclasses
import time
import datetime
import os, csv

from .XLF_FormGenerator import generate_controls
from .XLC_Excel_Consts import *
from .P01_Worksheetcalc import Calc_Worksheet
from pathlib import Path
import subprocess
import proggen.M09_Translate as M09_Translate
import proggen.M20_PageEvents_a_Functions as M20
import proggen.M27_Sheet_Icons as M27


try:
    keyboard_Module_imported = False
    import keyboard
    keyboard_Module_imported = True
    logging.debug("Keyboard Module imported")
except: 
    logging.debug("Keyboard Module not imported")

pyProgfile_dir = "\\LEDs_AutoProg\\pyProg_Generator_MobaLedLib"

shift_key = False

guifactor  = 1.55

global_worksheet_dict = {}

global_controller = None

Now = 0

def checkplatform(checkstr):
    teststr = platform.system()
    logging.debug("Checkplatform: Current platform="+teststr+"- Checked Platform="+checkstr)
    return teststr==checkstr

def get_home_dir():
    logging.debug("get_home_dir")
    homedir = pathlib.Path.home()
    logging.debug("HomeDir:"+homedir)
    return homedir

def Center_Form(form):
    return

def updateWindow():
    global_controller.update()

def getvalue(row,column):
    value=ActiveSheet.tksheet.get_data(row-1,column-1)
    #print("P01.getvalue:",row,column,value)
    return value

def activate_workbook(workbookname):
    #find workbook
    for workbook in Workbooks:
        if workbook.Name==workbookname:
            set_activeworkbook(workbook)
            return
    return

def set_activeworkbook(p_workbook):
    global ActiveSheet,ActiveWorkbook,Selection,Worksheets
    Debug.Print("Set_activeworkbook:", p_workbook.Name)
    aw_Name="---"
    as_Name="---"
    naw_Name="---"
    nas_Name="---"
    if ActiveWorkbook!=p_workbook:
        
        if ActiveWorkbook:
            aw_Name=ActiveWorkbook.Name
        if ActiveSheet:
            as_Name=ActiveSheet.Name
        if p_workbook:
            naw_Name=p_workbook.Name
            if p_workbook.ActiveSheet:
                nas_Name=p_workbook.ActiveSheet.Name
        #print("Set_activeworkbook:",aw_Name,as_Name,"->",naw_Name, nas_Name)
        if ActiveWorkbook!= None:
            ActiveWorkbook.ActiveSheet = ActiveSheet # save active sheet of current workbook
            ActiveWorkbook.Selection = Selection # save selection in workbook
        ActiveWorkbook = Workbook = p_workbook
        if p_workbook.ActiveSheet == None:
            ActiveSheet = ActiveWorkbook.Sheets(p_workbook.start_sheet)
            p_workbook.ActiveSheet = ActiveSheet
        else:
            ActiveSheet = p_workbook.ActiveSheet
        #print("Set_activeworkbook:",p_workbook.Name, p_workbook.ActiveSheet.Name)
        global_controller = p_workbook.controller
        global_controller.activeworkbook = p_workbook
        if p_workbook.Selection == None:
            Selection = CSelection(None,ws=p_workbook.ActiveSheet)
        else:
            Selection = p_workbook.Selection
        Worksheets = p_workbook.sheets
        if p_workbook.ActiveSheet != None:
            p_workbook.ActiveSheet.Activate()

def val(value):
    valtype=type(value)
    if valtype is CRange:
        cellvalue = str(value.get_value())
        if cellvalue != "":
            if "§" in cellvalue:
                cellpart = cellvalue.split("§")
                cellvalue = cellpart[0]
            if IsNumeric(cellvalue):
                return int(float(cellvalue))
            else:
                return 0
        else:
            return 0
    elif valtype is str:
        if value != "":
            if IsNumeric(value):
                try:
                    return int(value)
                except:
                    return int(float(value))
            else:
                return 0
        else:
            return 0
    return int(value)
    
def set_statusmessage(message, monitor_message=False):
    global_controller.set_statusmessage(message, monitor_message=monitor_message)
    global_controller.update()

__VK_LBUTTON = 0x1
__VK_RBUTTON = 0x2
__VK_MBUTTON = 0x4
__VK_UP = 0x26
__VK_DOWN = 0x28
__VK_RETURN = 0xD
__VK_ESCAPE = 0x1B
__VK_CONTROL = 0x11
__VK_SHIFT = 0x10
VK_SHIFT = 0x10

def GetCursorPos():
    pass

def SetCursorPos(x,y):
    pass

class Worksheet(object):
    """Placeholder for Worksheettype"""
    def __init__(self):
        self.Name = "Dummy"

# https://learn.microsoft.com/en-us/office/vba/api/excel.workbook

class CWorkbook(object):
    def __init__(self, frame=None,path=None,pyProgPath=None,workbookName=None, workbookFilename=None, workbookdict=None,init_workbook=None,open_workbook=None,start_sheet=None,controller=None,width=None,height=None, application=None):
        global Workbooks,pyProgfile_dir,Worksheets, DefaultFont
        # Row and Columns are 0 based and not 1 based as in Excel
        if controller==None:
            self.controller = global_controller
        else:
            self.controller = controller
        if application == None:
            self.Application = ActiveApplication
        else:
            self.Application = application
        self.controller.set_statusmessage("Erstelle Workbook: "+workbookName)
        #self.controller.bind("<KeyPress-Shift_L>", self.shift_press)
        #self.controller.bind("<KeyPress-Shift_R>", self.shift_press)
        #self.controller.bind("<KeyRelease-Shift_L>", self.shift_release)
        #self.controller.bind("<KeyRelease-Shift_R>", self.shift_release)
        self.Application = application

        if workbookName:
            self.Name = workbookName
        else:
            self.Name = "PyProgramWorkbook"
        Workbooks.append(self)
        self.workbookfilename = None
        if frame != None:
            self.Path = path
            self.pyProgPath = pyProgPath
            self.master = frame
            self.init_workbook_proc=init_workbook
            self.open_workbook_proc = open_workbook
            self.tabframedict = {}
            self.oldTabName=""
            self.ActiveSheet = None
            self.workbook_width  = width
            self.workbook_height = height
            self.default_font = self.controller.defaultfontnormal
            DefaultFont = self.default_font
            self.FullName = workbookFilename
            self.ws_filename = None
            self.wb = None
            style = ttk.Style(frame)
            style.configure('downtab.TNotebook', tabposition='sw')
            self.container = ttk.Notebook(frame, style="downtab.TNotebook")
            self.container.bind("<Alt-B1-Motion>", self.reorder_tabs)
            self.container.grid_rowconfigure(0,weight=1)
            self.container.grid_columnconfigure(0,weight=1)                 
            self.container.grid(row=0,column=0,columnspan=2,sticky="nesw")
            self.tabdict = dict()
            self.tabdict_frames = dict()
            self.sheets = list()
            self.workbookdict = workbookdict
            self.sheetdict = workbookdict.get("SheetDict", {})
            self.jumptable = workbookdict.get("JumpTable", {}) 
            self.SheetsDict = CSheetsDict()
            self.Selection = None
            self.tabid    = 0
            self.Events_dict = self.workbookdict.get("Events",None)
            for sheetname in self.sheetdict.keys():
                self.add_sheet(sheetname)
            self.container.bind("<<NotebookTabChanged>>",self.TabChanged)
        else:
            self.Path = path
            self.tksheet= None
            self.sheets = None
        self.start_sheet=start_sheet
        set_activeworkbook(self)
        self.InitWorkbookFunction=None
        #Worksheets = self.sheets

        
    def reorder_tabs(self,event):
        try:
            index = self.container.index(f"@{event.x},{event.y}")
            self.container.insert(index, child=self.container.select())
    
        except tk.TclError:
            pass    
        
    def shift_press(self,event):
        global shift_key
        #print("Shift press")
        shift_key=True
    
    def shift_release(sekf,event):
        global shift_key
        #print("Shift release") 
        shift_key=False
        
    def init_workbook(self):
        self.controller.set_statusmessage("Init Workbook: "+self.Name)
        #add controls to sheet
        for sheet in self.sheets:
            self.controller.set_statusmessage("Init Workbook: "+self.Name + " Sheet: "+sheet.Name)
            sheetname_prop = self.sheetdict.get(sheet.Name, {})
            controls_dict = sheetname_prop.get("Controls",None)
            if controls_dict:
                sheet.add_controls(controls_dict)
            sheet.tksheet.redraw()
            sheet.drawShapes()
        #self.controller.bind("<<SheetChange>>",self.Evt_SheetChange)
        #self.controller.bind("<<SheetSelectionChange>>",self.Evt_SheetSelectionChange)
        #self.controller.bind("<<SheetReturnKey>>",self.Evt_SheetReturnKey)
        set_activeworkbook(self)
        self.activate_sheet(self.start_sheet)
        if self.InitWorkbookFunction:
            self.controller.set_statusmessage("Init WorkbookFunction: "+self.Name)
            self.InitWorkbookFunction(self)        

    def open(self):
        self.Evt_Workbook_Open()
        #if self.Name != "PatternWorkbook":
        #    if M28.Get_Num_Config_Var_Range("SimAutostart", 0, 3, 0) == 1: #               ' 04.04.22: Juergen Simulator
        #        M39.OpenSimulator()
        
    def Delay(self,dSeconds):
        #print("Delay started", dSeconds)
        self.controller.after(int(dSeconds*1000))
        #print("Delay End")
        
        
    def delete_sheet(self,sheetname):
        #remove sheet from all lists:
           
        #self.sheets.append(act_worksheet)
        #self.container.add(tabframe, text=sheetname)
        #self.tabframedict[sheetname]=act_worksheet
        logging.debug("Delete sheet:"+sheetname)
        tabframe=self.tabframedict[sheetname]
        curr_tabid = tabframe.sheet.tabid
        self.sheets.remove(tabframe.sheet)
        del self.tabframedict[sheetname]
        for sheet in self.sheets:
            if sheet.tabid > curr_tabid:
                sheet.tabid -= 1
        self.tabid -= 1
        self.container.forget(tabframe)
        return
    
    def add_sheet(self,sheetname ,from_sheet=None,After=None,nocontrols=False,noredraw=False):
        tabframe = ttk.Frame(self.container,relief="ridge", borderwidth=1)
        act_worksheet = self.new_sheet(sheetname,tabframe,from_sheet=from_sheet)
        tabframe.sheetname = sheetname
        tabframe.sheet = act_worksheet
        self.sheets.append(act_worksheet)
        self.container.add(tabframe, text=sheetname)
        self.tabframedict[sheetname]=tabframe
        if from_sheet==None:
            copyfrom_sheet=None
        else:
            copyfrom_sheet=self.Sheets(from_sheet)
        if copyfrom_sheet:
            act_worksheet.wscalculation_callback = copyfrom_sheet.wscalculation_callback
            act_worksheet.wschanged_callback = copyfrom_sheet.wschanged_callback
            act_worksheet.wsselected_callback = copyfrom_sheet.wsselected_callback
            col_pos = copyfrom_sheet.tksheet.get_column_widths(canvas_positions=True)
            if col_pos != []:
                act_worksheet.tksheet.set_column_widths(column_widths=col_pos, canvas_positions=True)            
            if not nocontrols:
                sheetname_prop = self.sheetdict.get(copyfrom_sheet.Name)
                if sheetname_prop == None:
                    sheetname_prop = copyfrom_sheet.sheet_property_dict
                controls_dict = sheetname_prop.get("Controls",None)
                if controls_dict:
                    act_worksheet.add_controls(controls_dict)
            if not noredraw:
                act_worksheet.tksheet.redraw()

        if not self.showws and not self.controller.show_hiddentables:
            act_worksheet.Visible = False
            #self.container.tab(tabframe,state="hidden")
        self.tabid += 1
        act_worksheet.Redraw_table()

    def new_sheet(self,sheetname,tabframe: tk.Frame,from_sheet=None) -> Object:
        if from_sheet==None:
            sheet_property_dict = self.sheetdict.get(sheetname)
        else:
            sheet_property_dict = self.sheetdict.get(from_sheet)
        if sheet_property_dict != None:
            sheettype = sheet_property_dict.get("SheetType","")
        else:
            sheettype = "DCC"
            sheet_property_dict = self.sheetdict.get(sheettype)
        self.showws = sheettype in ["Datasheet","Config"]
        filename=sheet_property_dict.get("Filename",None)
        self.fieldnames=sheet_property_dict.get("Fieldnames",None)
        self.RefreshShapesafterLoad =  sheet_property_dict.get("RefreshShapesafterLoad",False)
        if filename==None:
            filepathname = self.ws_filename # take filename from first worksheet loaded (for excel workbooks)
        else:
            filepathname = Path(self.pyProgPath) / filename
        act_worksheet = CWorksheet(sheetname,workbook=self, csv_filepathname=filepathname,frame=tabframe,sheet_property_dict=sheet_property_dict, tabid=self.tabid,width=self.workbook_width,height=self.workbook_height)
        return act_worksheet
    
    def activate_sheet(self,sheetname):
        sheet=self.Application.Worksheets.getWorksheetbyName(sheetname)
        self.container.select(sheet.Tabframe)
            
    def TabChanged(self,_event=None):
        newtab_name = self.container.select()
        if newtab_name != "":
            newtab = self.container.nametowidget(newtab_name)
            sheet_name = newtab.sheetname
            sheet = newtab.sheet
            if sheet:
                sheet.Activate()
                #print("Tabchanged - Sheet activated:",sheet_name)
            
            #newtab.tabselected()
        logging.debug("TabChanged %s - %s",self.oldTabName,newtab_name)        
        
    def Sheets(self,name=None) -> Object:
        if name==None:
            return self.sheets        
        if self.sheets != None:
            if name == "<<LASTSHEET>>":
                return self.sheets[-1]
            for sheet in self.sheets:
                if sheet.Name == name:
                    return sheet
        return None
    
    def Worksheets(self,name):
        if self.sheets != None:
            for sheet in self.sheets:
                if sheet.Name == name:
                    return sheet
        return None
    
    def SetInitWorkbookFunction(self,func):
        self.InitWorkbookFunction=func
    
    def Activate(self):
        set_activeworkbook(self)
        return
    
    def notimplemented(self,command):
        n = tk.messagebox.showinfo(command,
                               "Not implemented yet",
                                parent=self.master)
    
    def Save(self,filename=None):
        if filename == None:
            filename = tk.filedialog.asksaveasfilename(parent=self.master,
                                                        defaultextension='.json',
                                                        #initialdir=os.getcwd(),
                                                        filetypes=[("JSON","*.json"),
                                                          ("All files","*.*")])
        if filename:
            self.SaveWorkbook(filename)
        return
    
    def Load(self,filename=None):
        """load from a file"""
        if filename == None:
            filename = tk.filedialog.askopenfilename(parent=self.master,
                                                      defaultextension='.json',
                                                      #initialdir=os.getcwd(),
                                                      filetypes=[("JSON","*.json"),
                                                        ("All files","*.*")])
        if not os.path.exists(filename):
            #print ('file does not exist')
            return
        if filename:
            self.LoadWorkbook(filename,workbookname=self.Name, keep_filename=False)
        return
    
    def refreshicons(self):
        ActiveSheet.refreshicons()    
    """
    def is_datasheet(self,ws):
        _ret = False
        if self.Name=="PatternWorkbook":
            cellA1 = ws["A1"].value
            if cellA1 == "Normal Data Sheet":
                _ret = True
            else:
                _ret = False
        else:
            cellB1 = ws["B1"].value
            if cellB1!=None:
                if ( InStr(M02.AllData_PgIDs, ' ' + cellB1 + ' ') > 0): 
                    _ret = True
            else:
                _ret=False
        return _ret
    


    def LoadExcelWorkbook(self,filename=None):
        #load Excelworkbook from a file
        if filename == None:
            filename = tk.filedialog.askopenfilename(parent=self.master,
                                                      defaultextension='.xlsm',
                                                      #initialdir=os.getcwd(),
                                                      filetypes=[(self.Name+"-Excel","*.xlsm"),
                                                        ("All files","*.*")])
        if not os.path.exists(filename):
            #print ('file does not exist')
            return
        if filename:
            if self.ws_filename != filename:
                self.wb = openpyxl.load_workbook(filename)
                self.ws_filename = filename
            sheetnames=self.wb.sheetnames
            #print("Sheets:",sheetnames)
            for sheetname in sheetnames:
                ws = self.wb[sheetname]
                if self.is_datasheet(ws):
                    #print("Workbook:",self.Name)
                    #load this sheet
                    for row in ws.rows:
                        for cell in row:
                            if cell.value!=None:
                                #print(cell,cell.value,cell.comment)
                                #print(f'Row {cell.row}, Col {cell.column} = {cell.value}')
                                #print(f'{cell.value=} is at {cell.coordinate=}')                    
                                if cell.comment:
                                    pass # print(cell.comment.text)
                    named_range=self.wb.defined_names
                    #print(named_range)
            # the destinations attribute contains a list of ranges in the definitions
                       
        return    
    """
    def LoadSheetfromExcelWorkbook(self,worksheet, filename, fieldnames):
        """load Excelworkbook from a file"""
        if not os.path.exists(filename):
            logging.debug ('file does not exist'+filename)
            return
        if filename:
            if self.ws_filename != filename:
                self.wb = openpyxl.load_workbook(filename)
                self.ws_filename = filename
            if fieldnames == None:
                fieldnames = self.fieldnames
            if worksheet.Name != "Named_Ranges":
                if worksheet.Name in self.wb.sheetnames:
                    xlws = self.wb[worksheet.Name]
                else:
                    xlws = self.wb["DCC"]
                dictdata = []
                count=0                
                for row in xlws.rows:
                    rec_dict = []
                    #rec_dict = {key: "" for key in fieldnames} # create empty row with fieldnames
                    for cell in row:
                        if cell.value!=None:
                            if cell.data_type != "f": # no formulas
                                #print(cell,cell.value,cell.comment)
                                text=str(cell.value)
                                if text == chr(252) or text == chr(61692):
                                     # old hook value
                                    text = chr(xlhookchar)
                                if len(text)>0 and text[0]=="'": # remove leading '
                                    text=text[1:]
                                    #print(cell,text,cell.comment)
                                if cell.comment:
                                    #print(cell.comment.text)
                                    text=text+"§"+cell.comment.text
                                if True: #cell.column < len(fieldnames):
                                    rec_dict.append(text)
                                else:
                                    rec_dict.append(text)
                                #print(f'Row {cell.row}, Col {cell.column} = {cell.value}')
                                #print(f'{cell.value=} is at {cell.coordinate=}')
                            else:
                                rec_dict.append("")
                                #print("Formula:", cell.value)
                                pass
                        else:
                            rec_dict.append("")
                    dictdata.append(rec_dict)
                    count=count+1
                #worksheet.table.model = worksheet.table #TableModel()
                #worksheet.table.model.importDict(dictdata)
                worksheet.tksheet.set_sheet_data(data = dictdata)
                worksheet.tksheet.shapelist=[]                
                worksheet.tksheet.modelname=filename                            
            else:
                named_ranges=self.wb.defined_names
                dictdata = []
                count = 0
                
                for key in named_ranges.keys():
                    rec_dict = []
                    named_range = named_ranges[key]
                    #rec_dict = {key: "" for key in fieldnames} # create empty row with fieldnames
                    rec_dict.append(named_range.name)
                    attributes = named_range.attr_text
                    attr_split=attributes.split("!")
                    if len(attr_split)!=2:
                        break                    
                    rec_dict.append(attr_split[0]) # sheetname
                    colrow_str = attr_split[1]
                    colrow_split=colrow_str.split("$")
                    column_str = colrow_split[1]
                    column = ord(column_str[0])-ord("A")+1
                    row_str    = colrow_split[2]
                    row    = int(row_str)
                    rec_dict.append(column)
                    rec_dict.append(row)
                    dictdata.append(rec_dict)
                    count=count+1
                #worksheet.table.model = TableModel()
                #print(repr(dictdata))
                worksheet.tksheet.data = dictdata
                worksheet.tksheet.shapelist=[]                
                worksheet.tksheet.modelname=filename                      
            # the destinations attribute contains a list of ranges in the definitions
        return        
    
    def SaveWorkbook(self,filename,allsheets=False):
        workbookdata = {}
        for sheet in self.sheets:
            if sheet.Datasheet or sheet.Configsheet or allsheets:
                sheetdata = sheet.getData()
                workbookdata[sheet.Name] = copy.deepcopy(sheetdata)
        #fd = open(filename,'wb')
        try:
            saveData={}
            saveData["Version"]=self.controller.ProgVersion
            saveData["DataVersion"] = self.controller.DataVersion
            saveData["Workbookdata"]= copy.deepcopy(workbookdata)
            with open(filename, 'w', encoding='utf8') as outfile:
                json.dump(saveData, outfile, ensure_ascii=False, indent=4)
            logging.debug("Save Workbook:" + filename)            
            #pickle.dump(saveData,fd)
            #fd.close()
        except BaseException as e:
            logging.debug(e,  exc_info=True)            
            logging.debug("Workbook: Save Workbook Error: "+filename)            
        return
    
    def LoadWorkbook(self,filename: str,workbookname="", keep_filename=True):
        if shift_key:
            logging.debug("Load Workbook: Shift Key pressed:"+filename+" not loaded")
            return
        if not self.controller.loaddatafile:
            logging.debug("Load Workbook: arg loaddatafile=False - "+filename+" not loaded")
            return
        jsondata={}
        file_not_found=True
        try:
            try:
                with open(filename, "r", encoding='utf8') as read_file:
                    file_not_found=False
                    jsondata = json.load(read_file)
                    if keep_filename:
                        self.workbookfilename = filename
            except ValueError as err:
                logging.error ("ERROR: JSON Error in Config File %s",filename)
                logging.error(err)
            except:
                if file_not_found:
                    logging.warning ("Warning: Config File %s not found",filename)
                else:
                    logging.error ("ERROR: JSON Error in Config File %s",filename)
                    logging.error(jsondata)
                    jsondata = {}                
            #savedData = pickle.load(read_file)
            savedData = jsondata
            logging.debug("Load Workbook:"+filename)
            logging.debug("FileVersion:"+savedData.get("Version","000"))
            dataversion = savedData.get("DataVersion","000")
            if workbookname=="ProgGenerator":
                from_sheet = "DCC"
                if int(dataversion) < int(self.controller.ProgGenMinDataVersion):
                    return
            elif workbookname =="PatternWorkbook":
                from_sheet = "Main"
                if int(dataversion) < int(self.controller.PattGenMinDataVersion):
                    return                    
            workbookdata=savedData["Workbookdata"]
            for sheetname in workbookdata.keys():
                sheet = self.Sheets(sheetname)
                #print("Load WB checking sheets: "+sheetname)
                if sheet == None:
                    self.add_sheet(sheetname, from_sheet=from_sheet, nocontrols=True,noredraw=True)
            for sheet in self.sheets:
                data = copy.deepcopy(workbookdata.get(sheet.Name,{}))
                #print("Load WB Loading:"+sheet.Name)
                self.controller.set_statusmessage("Workbook: "+workbookname+" Loading:"+sheet.Name)
                self.controller.update()
                if data !={}:
                    sheet.setData(data)
                    #sheet.EventWScalculate(Cells(5,5))
                    Application.Caller="Update"
                    #sheet.EventWSchanged(Cells(5,5))
                    #self.Evt_Redraw_Sheet(sheet)
                    if False: #sheet.RefreshShapesafterLoad:
                        sheet.refreshicons()
                    else:
                        pass #sheet.EventWSchanged(Cells(5,5))  # testHL
        except BaseException as e:
            logging.debug(e, exc_info=True)            
            logging.debug("Workbook: Load Workbook Error"+filename)
        return
    
    def install_Betatest(self):
        self.notimplemented("install_Betatest")
    
    def library_status(self):
        self.notimplemented("library_status")
    
    def install_fast_bootloader(self):
        self.notimplemented("install_fast_bootloader")
        
    def start_patternconf(self):
        self.notimplemented("start_patternconf")
    
    def send_to_patternconf(self):
        self.notimplemented("send_to_patternconf")
    
    def receiver_from_patternconf(self):
        self.notimplemented("receiver_from_patternconf")
        
    def check_Data_Changed(self):
        for sheet in self.sheets:
            if sheet.check_Data_Changed():
                return True
        return False
    
    # ------------------------------------------
    # VBA events
    # ------------------------------------------
    def Noneproc(*args):
        pass
    
    def Call_EventProc(self,eventname):
        if self.Events_dict:
            eventproc = self.Events_dict.get(eventname,None)
            if eventproc:
                return eventproc
            else:
                return self.Noneproc
        else:
            return self.Noneproc
    
    def Evt_Workbook_Open(self):
        # occurs when the workbook is opened
        if Application.EnableEvents:
            self.Call_EventProc("Open")()
    
    def Evt_Workbook_BeforeClose(self):
        if Application.EnableEvents:
            self.Call_EventProc("BeforeClose")
            return
    
    def Evt_SheetTableUpdate(self,sheet,target):
        # Occurs after the sheet table has been updated.
        if Application.EnableEvents:
            self.Call_EventProc("SheetTableUpdate")(sheet,target)
    
    def Evt_SheetCalculate(self, sheet, target):
        # Occurs after any worksheet is recalculated or after any changed data is plotted on a chart.
        if Application.EnableEvents:
            Application.EnableEvents=False
            self.Call_EventProc("SheetCalculate")(sheet, target)
            Application.EnableEvents=True
    
    def Evt_SheetBeforeDoubleClick(self, sheet,target, cancel):
        # Occurs when any worksheet is double-clicked, before the default double-click action.
        if Application.EnableEvents:
            self.Call_EventProc("SheetBeforeDoubleClick")(sheet, target, cancel)
    
    def Evt_SheetChange(self, tks_event):
        if Application.EnableEvents:
            # Occurs when cells in any worksheet are changed by the user or by an external link.
            row=tks_event.selected.row+1
            col=tks_event.selected.column+1
            sheet = global_worksheet_dict[tks_event.sheetname]
            target=CRange(row,col)
            self.Evt_SheetCalculate(sheet,target)
            self.Call_EventProc("SheetChange")(target)
            sheet.Redraw_table()
        
    def Evt_SheetSelectionChange(self,tks_event):
        if Application.EnableEvents:
            Application.EnableEvents=False
            if tks_event.selected != ():
                row=tks_event.selected.row+1
                col=tks_event.selected.column+1
                #sheet = global_worksheet_dict[tks_event.sheetname]
                #sheet=widget.worksheet
                target=CRange(row,col)
                self.Call_EventProc("SheetSelectionChange")(target)
            Application.EnableEvents=True
        
    #def Evt_SheetActivate(self, tks_event):
    #    if Application.EnableEvents:
    #        sheet = global_worksheet_dict[tks_event.sheetname]
    #        self.Call_EventProc("SheetActivate")(sheet)
        
    #def Evt_SheetDeactivate(self, tks_event):
    #    sheet = global_worksheet_dict[tks_event.sheetname]
    #    if Application.EnableEvents:
    #        self.Call_EventProc("SheetDectivate")(sheet)                
    
    #def Evt_NewSheet(self, event):
    #    # Occurs when a new sheet is created in the workbook.
    #    sheet = global_worksheet_dict[tks_event.sheetname]
    #    if Application.EnableEvents:
    #        self.Call_EventProc("NewSheet")(sheet)
            
    def Evt_SheetReturnKey(self, tks_event):
        # Occurs when a new sheet is created in the workbook.
        if Application.EnableEvents:
            self.Call_EventProc("ReturnKey")()
    
    #def Evt_SheetOpen(self, event):
    #    # Occurs when a new sheet is opened
    #    if Application.EnableEvents:
    #        self.Call_EventProc("SheetOpen")()

    def Synch_Evt_SheetChange(self, widget):
        #return #test
        if Application.EnableEvents:
            # Occurs when cells in any worksheet are changed by the user or by an external link.
            row=widget.event_row+1
            col=widget.event_col+1
            sheet = global_worksheet_dict[widget.name]
            target=CRange(row,col,ws=sheet)
            #pattgen.M02_Main.Global_On_Enter_Proc()
            self.Evt_SheetCalculate(sheet,target)
            #Debug.Print("Synch_Evt_SheetChange: <<SheetChange>> "+str(row),","+str(col))
            sheet.Call_EventProc("Change")(target)
            sheet.sheet_was_modified(None)
            sheet.Redraw_table()
            
    def Evt_Redraw_Sheet(self, sheet):
        # occurs when the workbook is redrawn
        if Application.EnableEvents:
            sheet.Call_EventProc("Redraw")(sheet)

#***************************************************
#* class: CWorkbooks
#* Description: Implements the Worksbooks property of Application
#* https://learn.microsoft.com/en-us/office/vba/api/excel.workbooks
#***************************************************
class CWorkbooks(object):
    def __init__(self,application=None):
        if application==None:
            application = ActiveApplication
        self.Application = application
        self.currentworkbook=None
        self.currentidx = 0

    def _getWorkbookbyName(self, workbookname):
        for workbook in self.Application.workbookslist:
            if workbook.Name == workbookname:
                return workbook
        return None
        
    def getWorkbook(self,index):
        if type(index)==int:
            if index <= len(self.Application.workbookslist):
                workbook = self.Application.workbookslist[index-1]
            else:
                workbook=None
        else:
            workbook=self._getWorkbookbyName(index) # by name
        return workbook
    
    def getlist(self):
        return self.Application.workbookslist
    
    def Add(self, templateFilename):
        # creates a new workbook
        pass
    
    # for compatibility reasons
    def append(self, workbook):
        testwb = self._getWorkbookbyName(workbook.Name)
        if testwb == None:
            self.Application.workbookslist.append(workbook)
    
    def Close(self):
        pass

    def Count(self):
        return len(self.Application.workbooklist)
    
    def Open(self, ParentFrame = None, FileName=None, Format=None, Delimiter=",", path=None, pyProgPath=None, workbookName=None, workbookdict=None, init_workbook=None, open_workbook=None, start_sheet=None, controller=None, width=None, height=None, macro_caller=None, InitWorkBookFunction=None):
        global ThisWorkbook
        # open workbook and create all sheets and init workbook
        workbook = self._getWorkbookbyName(workbookName)
        if workbook == None:
            workbook = CWorkbook(frame=ParentFrame, path=path, pyProgPath=pyProgPath, workbookName=workbookName, workbookFilename=None, workbookdict=workbookdict, init_workbook=init_workbook, open_workbook=open_workbook, start_sheet=start_sheet, controller=controller, width=width, height=height)
        ThisWorkbook = workbook
        macro_caller.ThisWorkbook = workbook        
        workbook.LoadWorkbook(FileName,workbookname=workbook.Name)
        workbook.Application = self.Application
        if InitWorkBookFunction:
            workbook.SetInitWorkbookFunction(InitWorkBookFunction)
        workbook.init_workbook()  # test 15.4.2024
        workbook.open()
        return workbook

    def __iter__(self):
    #returning __iter__ object
        self.currentidx = 1
        self.currentworkbook = self.getWorkbook(self.currentidx)
        self.stopiteration=False
        return self
    
    def __next__(self):
        self.currentworkbook = self.getWorkbook(self.currentidx)
        if self.stopiteration or self.currentworkbook==None:
            raise StopIteration 
        self.currentidx += 1
        if self.currentidx > len(self.Application.workbookslist):
            self.stopiteration=True       
        return self.currentworkbook
    
    #*****************************
    #* Properties
    #*****************************
    def _set_count(self, value):
        pass
    def _get_count(self):
        return len(self.Application.workbooklist)
    Count = property (_get_count, _set_count, doc="Count")

#***************************************************
#* class: CWorksheets
#* Description: Implements the Workssheets property of Application
#* https://learn.microsoft.com/en-us/office/vba/api/excel.application.worksheets
#***************************************************
class CWorksheets(object):
    def __init__(self,application=None):
        if application==None:
            application = ActiveApplication
        self.Application = application
        self.currentworksheet = None
        self.currentidx = 0

    def getWorksheetbyName(self, worksheetname):
        for worksheet in self.Application.ActiveWorkbook.sheets:
            if worksheet.Name == worksheetname:
                return worksheet
        return None
        
    def getWorksheet(self,index):
        if type(index)==int:
            if index <= len(self.Application.ActiveWorkbook.sheets):
                worksheet = self.Application.ActiveWorkbook.sheets[index-1]
            else:
                worksheet=None
        else:
            worksheet=self.getWorksheetbyName(index) # by name
        return worksheet
    
    def getlist(self):
        return self.Application.ActiveWorkbook.sheets
    
    def Add(self, Before=None, After=None, Count=1, Type=xlWorksheet):
        pass
    
    def append(self, worksheet):
        self.Application.ActiveWorkbook.sheets.append(worksheet)
    
    def Copy(self, Before=None, After=None):
        pass

    def Count(self):
        return len(self.Application.ActiveWorkbook.sheets)
    
    def Delete(self):
        pass
    
    def Move(self, Before=None, After=None):
        pass
    
    def Select(self, Replace=True):
        pass
    
    def __iter__(self):
    #returning __iter__ object
        self.currentidx = 1
        self.currentworksheet = self.getWorksheet(self.currentidx)
        self.stopiteration=False
        return self
    
    def __next__(self):
        self.currentworksheet = self.getWorksheet(self.currentidx)
        if self.stopiteration or self.currentworksheet==None:
            raise StopIteration 
        self.currentidx += 1
        if self.currentidx > len(self.Application.ActiveWorkbook.sheets):
            self.stopiteration=True
        return self.currentworksheet
    
    #*****************************
    #* Properties
    #*****************************
    def _set_count(self, value):
        pass
    def _get_count(self):
        return len(self.Application.ActiveWorkbook.sheets)
    Count = property (_get_count, _set_count, doc="Count")
    
    
class CCellsFind(object):
    def __init__(self,ws=None):
        self.sheet=ws
    
    def Find(self,What="", After=None, LookIn=xlFormulas, LookAt= xlWhole, SearchOrder= xlByRows, SearchDirection= xlNext, MatchCase= True, SearchFormat= False):
        cells=self.sheet.Cells()
        if After!=None:
            cells.startrow=After.Row
            cells.startcol=After.Column
                
        for cell in cells:
            cell_val=str(cell.Value)
            if "§" in cell_val:
                special_char = "§" # remove comments in excel sheet
                index = cell_val.find(special_char)
                if index != -1:
                    cell_val = cell_val[:index]
            if cell_val == What:
                return cell
        return None
    
class CButtons(object):
    def __init__(self,ws=None):
        self.shape=None
        self.ws= ws
        self.Buttonlist = []

    def Add(self,l,t,w,h):
        shapetype = msoOLEControlObject
        Left = l
        Top = t
        Width = w
        Height = h
        control_dict = {  "Name"          : "Button",
                                     "BackColor"     : "#FFFFFF",
                                     "BorderColor"   : "#000012",
                                     "Caption"       : "Button",
                                     "Height"        : Height,
                                     "Left"          : Left,
                                     "Col"           : None,
                                     "Top"           : Top,
                                     "Row"           : None,                                        
                                     "Type"          : "FormWindow",
                                     "Visible"       : True,
                                     "Width"         : Width,
                                     "Components"    : [{"Name":"Button","Accelerator":"","BackColor":"#00000F","BorderColor":"#000006","BorderStyle":"fmBorderStyleNone","IconName":"",
                                                        "Caption":"Test",
                                                        "Command": None ,"ControlTipText":"","ForeColor":"#000012","Height":Height-2,"Left":1,"Top":1,"Type":"CommandButton","Visible":True,"Width":Width-2},
                                                        ]}
                        
        self.shape = self.ws.Shapes.AddShape(shapetype, Left, Top, Width, Height, Row=None, Col=None, Fill=0,text="",name="",control_dict=control_dict)
        self.Buttonlist.append(self.shape)
        return self.shape
       
class CWorksheet(object):
    
    def __init__(self,Name,workbook=None,csv_filepathname=None,frame=None,sheet_property_dict=None,tabid=0,width=None, height=None):
        
        sheetdict_popmenu = {"DCC":
                             { "LED blinken Ein/Aus":self.LED_flash,
                               "LED nacheinander blinken Ein/Aus":self.LED_flash_seq
                             }
                            }
        self.sheet_property_dict = sheet_property_dict 
        Sheettype = sheet_property_dict.get("SheetType","")
        callback = Sheettype in ["Datasheet","Libraries"]
        self.showws = Sheettype in ["Datasheet","Config"]
        self.formating_dict = sheet_property_dict.get("Formating",None)
        self.Events_dict = sheet_property_dict.get("Events", {})
        fieldnames = sheet_property_dict.get("Fieldnames",None)
        if type(fieldnames) == str:
            fieldnames = fieldnames.split(";")
        self.fieldnames = fieldnames
        self.saveShapes = sheet_property_dict.get("SaveShapes",True)
        self.RefreshShapesafterLoad =  sheet_property_dict.get("RefreshShapesafterLoad",False)
        
        screen_width = 1920
        screen_height = 1080
        self.controller = workbook.controller
        self.controller.set_statusmessage("Erstelle WorkSheet: "+Name)
        #print(workbook.controller.winfo_width(),workbook.controller.winfo_height())
        self.rightclickactiondict = sheetdict_popmenu.get(Name,{})
        
        if width == None:
            self.width = screen_width-120
        else:
            self.width = width
        if height == None:
            self.height = screen_height-350
        else:
            self.height= height
            
        self.Workbook = workbook
        self.worksheet = self
        self.default_font = workbook.default_font
        self.thefont = self.default_font #('Arial',9)
        self.ProtectContents = False
        self.Datasheet = Sheettype == "Datasheet"
        self.Configsheet = Sheettype == "Config"
        self.csv_filepathname = csv_filepathname
        self.Name = Name
        self.shape_clicked = False
        self.Tabframe = frame
        self.DataChanged=False
        self.wsRowdict = {}
        self.EnableCalculation=True
        self.tabid = tabid
        self.Buttons = CButtons(ws=self)
        self.shapelist = [] #*HL list of shapes
        self.imagedict = {}
        self.Flag_move=False
        self.lastselectedRows = []
        self.Gridlist = {}
        self.grid_color = "#000000"
        self.linewidth = 1
        self.wschanged_callback = None
        self.wscalculation_callback = None
        self.wsselected_callback = None        
        self._visible_val = True
        
        if False: #tablemodel:
            #self.tablemodel = tablemodel
            self.tksheet = Sheet(frame, name=self.Name, width=self.width, height=self.height, data=[[f"Row {r}, Column {c}\nnewline 1\nnewline 2" for c in range(20)] for r in range(21)]) #test TableCanvas(frame, tablename=Name, model=tablemodel,width=self.width,height=self.height,default_font=self.default_font,scrollregion=(0,0,self.width,self.height),rightclickactions=self.rightclickactiondict,worksheet=self)
            global_worksheet_dict[self.Name] = self
            self.tksheet.worksheet = self
            frame.grid_columnconfigure(0, weight=1)
            frame.grid_rowconfigure(0, weight=1)             
            self.tksheet.grid(row = 0, column = 0, sticky = "nswe")
            self.tksheet.enable_bindings()
            #self.tablemodel = self.tksheet #test
        else:
            if csv_filepathname:
                #self.tablemodel = TableModel()
                self.tksheet = Sheet(frame, name=self.Name, width=self.width, height=self.height, data=[[f"Row {r}, Column {c}\nnewline 1\nnewline 2" for c in range(20)] for r in range(21)]) #TableCanvas(frame, tablename=Name, model=None,width=self.width,height=self.height,default_font=self.default_font,scrollregion=(0,0,self.width,self.height),rightclickactions=self.rightclickactiondict,worksheet=self)
                global_worksheet_dict[self.Name] = self
                self.tksheet.worksheet = self
                self.tksheet.enable_bindings()
                self.update_table_properties()
                frame.grid_columnconfigure(0, weight=1)
                frame.grid_rowconfigure(0, weight=1)                 
                self.tksheet.grid(row = 0, column = 0, sticky = "nswe")
                fileextension = os.path.splitext(self.csv_filepathname)[1]
                fileextension = fileextension.upper()
                if fileextension == ".CSV":
                    self.importCSV(filename=self.csv_filepathname, sep=';',fieldnames=self.fieldnames)
                elif fileextension == ".MLL_PCF":
                    self.Workbook.load()
                elif fileextension == ".XLSM":
                    self.Workbook.LoadSheetfromExcelWorkbook(self,self.csv_filepathname, self.fieldnames)
                #self.tablemodel = self.tksheet #.getModel()
                if callback:
                    self.set_left_click_callback(self.left_click_callback) #test self.table.set_left_click_callback(self.left_click_callback)
            else: #test
                self.tksheet = Sheet(frame, name=self.Name, width=self.width, height=self.height, data=[[f"Row {r}, Column {c}\nnewline 1\nnewline 2" for c in range(20)] for r in range(21)]) #TableCanvas(frame, tablename=Name, model=None,width=self.width,height=self.height,default_font=self.default_font,scrollregion=(0,0,self.width,self.height),rightclickactions=self.rightclickactiondict,worksheet=self)
                global_worksheet_dict[self.Name] = self
                self.tksheet.worksheet = self
                frame.grid_columnconfigure(0, weight=1)
                frame.grid_rowconfigure(0, weight=1)                
                self.tksheet.grid(row = 0, column = 0, sticky = "nswe")
                self.tksheet.enable_bindings()
                #self.tablemodel = self.tksheet
                self.update_table_properties()
                return
        if self.formating_dict:
            self.handle_hidden_cells(self.formating_dict.get("HideCells",[]))
            self.handle_hidden_rows(self.formating_dict.get("HideRows",[]))
            self.handle_protected_cells(self.formating_dict.get("ProtectedCells",[]))
            self.handle_formating(self.formating_dict.get("FontColor",[]))
            self.handle_gridlist(self.formating_dict.get("GridList",[]))
            self.handle_gridformat(self.formating_dict.get("GridFormat",{}))
            self.handle_ColumnAlignment(self.formating_dict.get("ColumnAlignment",[]))
            self.handle_ColumnWidth(self.formating_dict.get("ColumnWidth",[]))

            
            #self.handle_events_dict(self.formating_dict)

            #test colsizelist = self.formating_dict.get("ColumnWidth",[])
            #test for col in range(0, len(colsizelist)):
            #test    self.table.column_width(column=col, width=colsizelist[col])
            #self.table.resizecolumns(colsizelist,redraw=True)
            

            rowheight=self.formating_dict.get("RowHeight",30)
            self.tksheet.set_all_row_heights(height=rowheight)

        self.update_table_properties()
        self.DataChanged=False
        #*HL dummy for testing
        self.CellsFind=CCellsFind(ws=self)
        self.Controls = []
        self.Controls_Dict = {}
        self.persistent_controls_dict = {}
        self.RangeDict = CRangeDict()
        self.wsInitDataFunction=None
        
    def init_data(self):
        if self.wsInitDataFunction:
            self.wsInitDataFunction(self)
        #F00.worksheet_init(self)
        self.DataChanged=False
        
    def sheet_was_modified(self, event: dict):
        self.DataChanged=True
        
    def update_table_properties(self):
        self.LastUsedRow_val = self.tksheet.get_total_rows() #getLastUsedRow()+1
        self.LastUsedColumn_val = self.tksheet.get_total_columns() # ColumnCount
        self.UsedRange_val = CRange((1,1) , (self.LastUsedRow_val,self.LastUsedColumn_val),ws=self)
        self.MaxRange_val  = CRange((1,1) , (self.LastUsedRow_val,self.LastUsedColumn_val),ws=self)
        #self.Rectangles = CRectangles()
        self.AutoFilterMode = False
        self.searchcache = {}
        self.Shapes = CShapeList(self)
        self.CellDict = CCellDict(ws=self)
        self.End_val = self.LastUsedColumn_val
        
        self.tksheet.extra_bindings("all_modfied_events", func=self.Evt_SheetChange)
        self.tksheet.extra_bindings("all_select_events", func=self.Evt_SheetSelectionChange)
        self.tksheet.extra_bindings("cell_select", func=self.EventSH_cellselected)
        self.tksheet.extra_bindings("row_height_resize", func=self.tkEvent_row_height_resize)
        self.tksheet.extra_bindings("column_width_resize", func=self.tkEvent_column_width_resize)
        #self.tksheet.extra_bindings ("edit_cell", func=self.Workbook.Synch_Evt_SheetReturnKey)
        self.tksheet.extra_bindings ("sheetmodified", func=self.Evt_SheetChange)
        #self.tksheet.extra_bindings ("paste", func=self.Evt_Sheet_End_Paste)
                
                
        self.tksheet.bind("<<SheetModified>>", self.sheet_was_modified)
        self.tksheet.bind("<Double-Button-1>", self.EventWSdoubleclick)
        #self.tksheet.bind("<<Paste>>", self.EventWSPaste, add="+")

        
        self.tksheet.popup_menu_add_command(
                    "Set Color",
                    self.setCellColor,
                    index_menu=False,
                    header_menu=False,
                    empty_space_menu=False)
        
        
    def setCellColor(self):
            current_selection = self.getSelectedCell()
            if current_selection:
                box = (current_selection[0] - 1, current_selection[1] - 1)
                try:
                    bgcolor = self.tksheet[box].bg
                except:
                    bgcolor = "#FFFFFF"
                clr = self.getaColor(bgcolor)
                if clr != None:
                    # set cell data, end user Undo enabled
                    # more information at:
                    # https://github.com/ragardner/tksheet/wiki/Version-7#setting-sheet-data
                    self.tksheet[box].bg = clr
                    
    def getaColor(self, oldcolor):
        ctuple, newcolor = tk.colorchooser.askcolor(title='pick a color', initialcolor=oldcolor,
                                                   parent=self.Tabframe)
        if ctuple == None:
            return None
        return str(newcolor)
    
    def Noneproc(*args):
        pass    
    
    def Call_EventProc(self,eventname):
        if self.Events_dict:
            eventproc = self.Events_dict.get(eventname,None)
            if eventproc:
                logging.debug("Call_EventProc "+eventname)
                return eventproc
            else:
                return self.Noneproc
        else:
            return self.Noneproc    
    
    def Evt_SheetTableUpdate(self,sheet,target):
        # Occurs after the sheet table has been updated.
        if Application.EnableEvents:
            logging.debug("Evt_SheetBeforeDoubleClick:"+repr(target))
            self.Call_EventProc("TableUpdate")(sheet,target)
    
    def Evt_SheetCalculate(self, sheet, target):
        # Occurs after any worksheet is recalculated or after any changed data is plotted on a chart.
        if Application.EnableEvents:
            logging.debug("Evt_SheetBeforeDoubleClick:"+repr(target))
            Application.EnableEvents=False
            Calc_Worksheet(sheet, target)
            self.Call_EventProc("Calculate")()
            Application.EnableEvents=True
    
    def Evt_SheetBeforeDoubleClick(self, sheet,target, cancel):
        # Occurs when any worksheet is double-clicked, before the default double-click action.
        if Application.EnableEvents:
            logging.debug("Evt_SheetBeforeDoubleClick:"+repr(target))
            self.Call_EventProc("BeforeDoubleClick")(target, cancel)
            self.Workbook.Call_EventProc("SheetBeforeDoubleClick")(sheet, target, cancel)
    
    def Evt_SheetChange(self, tks_event):
        if Application.EnableEvents:
            logging.debug("Evt_SheetChange:"+repr(tks_event))
            # Occurs when cells in any worksheet are changed by the user or by an external link.
            row=tks_event.selected.row+1
            col=tks_event.selected.column+1
            sheet = global_worksheet_dict[tks_event.sheetname]
            target=CRange(row,col)
            self.Evt_SheetCalculate(sheet,target)
            self.Call_EventProc("Change")(target)
            self.Workbook.Call_EventProc("SheetChange")(sheet, target)
            sheet.Redraw_table()
            
    def Evt_Sheet_End_Paste(self, tks_event):
        row = tks_event.selected.row
        s = tks_event.data[0][1]
        if s != '':
            M27.FindMacro_and_Add_Icon_and_Name(s, row, ActiveSheet)        
        M20.Update_Start_LedNr()
        
        
    def EventWSPaste(self, event):
        row,col = self.getSelectedCell()
        s = self.Cells(row, 10) #M25.Config__Col)
        if s != '':
            M27.FindMacro_and_Add_Icon_and_Name(s, Row, ActiveSheet)        
        M20.Update_Start_LedNr()
        
        
        
    def Evt_SheetSelectionChange(self,tks_event):
        if Application.EnableEvents:
            logging.debug("Evt_SheetSelectionChange:"+repr(tks_event))
            Application.EnableEvents=False
            if tks_event.selected != ():
                row=tks_event.selected.row+1
                col=tks_event.selected.column+1
                target=CRange(row,col)
                self.Call_EventProc("SelectionChange")(target)
                self.Workbook.Call_EventProc("SheetSelectionChange")(self, target)
            Application.EnableEvents=True
        
    def Evt_SheetActivate(self, tks_event):
        if Application.EnableEvents:
            logging.debug("Evt_SheetActivate:"+repr(tks_event))
            self.Call_EventProc("Activate")()
            self.Workbook.Call_EventProc("SheetActivate")(self)
        
    def Evt_SheetDeactivate(self, tks_event):
        if Application.EnableEvents:
            logging.debug("Evt_SheetDeactivate:"+repr(tks_event))
            self.Call_EventProc("Deactivate")()
            self.Workbook.Call_EventProc("SheetDeActivate")(self)
    
    def Evt_NewSheet(self, tks_event):
        # Occurs when a new sheet is created in the workbook.
        if Application.EnableEvents:
            logging.debug("Evt_NewSheet:"+repr(tks_event))
            self.Call_EventProc("NewSheet")()
            
    def Evt_SheetReturnKey(self, tks_event):
        # Occurs when a new sheet is created in the workbook.
        if Application.EnableEvents:
            logging.debug("Evt_SheetReturnKey:"+repr(tks_event))
            self.Call_EventProc("ReturnKey")()
    
    def Evt_SheetOpen(self, tks_event):
        # Occurs when a new sheet is opened
        if Application.EnableEvents:
            logging.debug("Evt_SheetOpen:"+repr(tks_event))
            self.Call_EventProc("Open")()

    def Synch_Evt_SheetChange(self, tks_event):
        #return #test
        if Application.EnableEvents:
            logging.debug("Synch_Evt_SheetChange:"+repr(tks_event))
            # Occurs when cells in any worksheet are changed by the user or by an external link.
            row=tks_event.event_row+1
            col=tks_event.event_col+1
            target=CRange(row,col,ws=self)
            #pattgen.M02_Main.Global_On_Enter_Proc()
            self.Evt_SheetCalculate(self,target)
            #Debug.Print("Synch_Evt_SheetChange: <<SheetChange>> "+str(row),","+str(col))
            self.Call_EventProc("Change")(target)
            self.sheet_was_modified(None)
            self.Redraw_table()
            
    def Evt_Redraw_Sheet(self, sheet):
        # occurs when the workbook is redrawn
        if Application.EnableEvents:
            logging.debug("Evt_Redraw_Sheet:"+sheet.Name)
            self.Call_EventProc("Redraw")(sheet)
    
        
    def EventSH_cellselected(self, event):
        #print("EventSH_cellselected:", event.selected)
        logging.debug("EventSH_cellselected")
        if self.Flag_move:
            cur_row = self.getSelectedRow()
            rows = self.lastselectedRows
            self.int_moveRows(rows,cur_row)
            self.Flag_move = False
        if self.left_click_callback:
            cur_row = self.getSelectedRow()
            cur_col = self.getSelectedColumn()
            if cur_col == None:
                logging.debug("EventSH_cellselected: CUR_COL==None "+repr(event))
            else:
                self.left_click_callback(cur_row, cur_col-1)
            
    def tkEvent_row_height_resize(self, event_data):
        #print(event_data.resized.rows)
        resized_rows = event_data.resized.rows
        logging.debug("tkEvent_row_height_resize:"+repr(resized_rows))
        for row in resized_rows.keys():
            row_data = resized_rows[row]
        self.moveShapesVertical(top_left_cell_row=row, deltaY=row_data["new_size"]-row_data["old_size"])
    
    def tkEvent_column_width_resize(self, event_data):
        #print(event_data.resized.columns)
        resized_columns = event_data.resized.columns
        logging.debug("tkEvent_column_width_resize:"+repr(resized_columns))
        for col in resized_columns.keys():
            col_data = resized_columns[col]
        self.moveShapesHorizontal(top_left_cell_column=col, deltaX=col_data["new_size"]-col_data["old_size"])
        
    def Calculate(self):
        self.EventWScalculate()
        
    def Delete(self):
        self.Workbook.delete_sheet(self.Name)
        
    def set_left_click_callback(self,callback):
        self.left_click_callback = callback
        
    def handle_hidden_cells(self, param):
        # not possible with TKSheet
        pass

    def handle_hidden_rows(self, param):
        if param != []:
            self.tksheet.hide_rows(param)
    
    def is_row_hidden(self, row):
        if ActiveSheet.tksheet.displayed_rows != []:
            return not row in ActiveSheet.tksheet.displayed_rows
        else:
            return False

    def handle_protected_cells(self, spanlist):
        for span in spanlist:
            self.tksheet.readonly(span)

    def handle_formating(self, param):
        for cellformat,formatdata in param.items():
            fg_color = formatdata["fg"]
            bg_color = formatdata["bg"]
            font = formatdata["font"]
            if len(font) == 2:
                font = (font[0], font[1], "normal")
            if cellformat != "default":
                celllist = formatdata["Cells"]
                for cell in celllist:
                    self.tksheet.highlight(cell, bg=bg_color, fg=fg_color, font=font)
            else:
                self.tksheet.set_options(font=font, header_font=font, index_font=font, table_bg=bg_color, table_fg=fg_color, table_grid_fg=fg_color, show_vertical_grid=False, show_horizontal_grid=False)

    def handle_gridlist(self, param):
        self.Gridlist = param
        
    def handle_gridformat(self, param):
        table_grid_fg = param.get("fg", "#808080")
        show_vertical_grid = param.get("show_vertical_grid", True)
        show_horizontal_grid = param.get("show_horizontal_grid", True)
        self.tksheet.set_options(table_grid_fg=table_grid_fg, show_vertical_grid=show_vertical_grid, show_horizontal_grid=show_horizontal_grid)

    def handle_ColumnAlignment(self, param):
        if isinstance(param, dict):
            for align, spanlist in param.items():
                for span in spanlist:
                    self.tksheet.align(span, align=align)
    
    def handle_ColumnWidth(self, param):
        colsizelist = param
        for col in range(0, len(colsizelist)):
            self.tksheet.column_width(column=col, width=colsizelist[col])
            
    #def handle_events_dict(self, formating_dict):
     #   Events_dict = formating_dict.get("Events", {})
     #   if Events_dict != {}:
     #       self.Events_dict = Events_dict
     #   else:
     #       self.Events_dict = {} #self.Workbook.Events_dict

        
    def add_controls(self,controls_dict):
        logging.debug("add_controls")
        for control_key in controls_dict.keys():
            if control_key !="Default":
                form_def = controls_dict[control_key]
                Left = form_def.get("Left",None)
                Top = form_def.get("Top",None)
                Row=form_def.get("Row",None)
                Col=form_def.get("Col",None)
                Width = form_def["Width"]
                Height = form_def["Height"]
                Name = form_def["Name"]
                logging.debug("add_controls:" +repr(form_def))
                controls_def= copy.deepcopy(form_def)
                self.Shapes.AddShape(msoOLEControlObject, Left, Top, Width, Height, Row=Row, Col=Col, Fill=0, text="", name=Name,control_dict=controls_def)
                
    def AddControl(self,control):
        self.Controls.append(control)
        self.Controls_Dict[control.Name]=control

    def set_control_val(self, control,value,disable=False):
        if value != None:
            control.Init_Value=value
            if hasattr(control,"TKVar"):
                variable = control.TKVar
                if variable:
                    if disable:
                        variable.config(state="normal") # we need to set the state to "normal" to be able to set values
                    try:
                        var_class = variable.winfo_class()
                    except: #limit_var has no winfo_class() function
                        var_class = "Limit_Var"                
                    if var_class == "TCombobox":
                        # value of combobox needs to be translated into a generic value
                        if isinstance(value,int):
                            current_index = value
                            value = variable.textvalues[current_index]
                        variable.set(value)
                    elif var_class in ["Entry"]:
                        variable.delete(0,tk.END)
                        variable.insert(0,value)
                    elif var_class in ["Text"]:
                        variable.delete(0.0,tk.END)
                        variable.insert(0.0,value)                          
                    else:          
                        variable.set(value)
                    if disable:
                        variable.config(state="disabled")                    
                else:
                    pass
        return    
            
    def get_control_val(self, control):
        value = ""
        variable = control.TKVar
        try:
            var_class = variable.winfo_class()
            if var_class == "TCombobox":
                # value of combobox needs to be translated into a generic value
                current_index = variable.current()
                if current_index == -1: # entry not in list
                    value = variable.get()
                else:
                    if variable.keyvalues !=[]:
                        value_list = variable.keyvalues
                        value = value_list[current_index]
                    else:
                        value=variable.get()
            elif var_class == "Text":
                value = variable.get("1.0",tk.END)            
            else:
                value = variable.get()
        except:
            value = variable.get()
        return value
        
    def SavePersistentControls(self):
        for control in self.Controls:
            if control.Persistent:
                self.persistent_controls_dict[control.Name]=self.get_control_val(control)
    
    def SetPersistentControls(self):
        self.tksheet.persistent_controls_dict = self.persistent_controls_dict
        for control in self.Controls:
            if control.Persistent:
                self.set_control_val(control,self.persistent_controls_dict[control.Name])
        
    def LED_flash(self,row,col):
        #print("LED Flash:",row,col)
        lednum_str = Cells(row+1,13)
        if lednum_str == "":
            lednum_str = Cells(row,13) # take LED num from previous line
            if lednum_str.isnumeric():
                lednum = int(lednum_str)
                ledcount_str = Cells(row,14)
                if ledcount_str.isnumeric():
                    ledcount=int(ledcount_str)
                else:
                    ledcount = 1
                lednum = lednum+ledcount+1 # next LED after current
                ledcount = 1
        else:
            if lednum_str.isnumeric():
                lednum = int(lednum_str)
                ledcount_str = Cells(row+1,14)
                if ledcount_str.isnumeric():
                    ledcount=int(ledcount_str)
                else:
                    ledcount = 1
            else:
                return
        global_controller.blinking_on_off(lednum,ledcount)
            
    def LED_flash_seq(self,row,col):
        #print("LED Flash_seq:",row,col)
        lednum_str = Cells(row+1,13)
        if lednum_str.isnumeric():
            lednum = int(lednum_str)
            ledcount_str = Cells(row+1,14)
            if ledcount_str.isnumeric():
                ledcount=int(ledcount_str)
            else:
                ledcount = 1
            global_controller.blinking_on_off(lednum,ledcount,seq=True)    
        
    def left_click_callback(self,value1,value2,callertype="cell"):
        if callertype=="cell":
            if self.shape_clicked:
                self.shape_clicked=False
                return False
            row=value1
            col=value2
            #print("Left_Click_Call_Back:",row,col)
            if False: #col == 1:  # aktive column
                cell = self.Cells(row+1,col+1)
                if cell.Value == "":
                    cell.Value = chr(xlhookchar)
                else:
                    cell.Value= ""
                self.Redraw_table()
                return False
            else:
                return True
        elif callertype=="canvas":
            #print("Leftclick-Canvas",value1)
            #callingshape = self.Shapes.shapelist[int(value1[0])]
            Application.caller = value1[0]
            if Application.canvas_leftclickcmd!=None:
                Application.canvas_leftclickcmd()
                self.shape_clicked = True # avoid cell maked as clicked
                return False
        
    def EnableMousePosition(self):
        pass
    
    def getSelectedRow(self):
        currently_selected = self.tksheet.get_selected_rows(get_cells_as_rows=True, return_tuple=True)
        if len(currently_selected) == 0:
            return 0
        else:
            return currently_selected[0]
        
    def getSelectedCell(self):
        currently_selected = list(self.tksheet.get_selected_cells())
        if currently_selected == []:
            return 1,1
        else:
            row = self.displayed_row_to_data(currently_selected[0][0]) + 1
            column = self.displayed_column_to_data(currently_selected[0][1]) + 1
            return row, column
        
    def getRowCount(self):
        rcount = self.tksheet.get_total_rows()
        return rcount
    
    def getColumnCount(self):
        ccount = self.tksheet.get_total_columns()
        return ccount 
        
    def getSelectedColumn(self):
        currently_selected = self.tksheet.get_currently_selected()
        if currently_selected:
            column = currently_selected.column
            type_ = currently_selected.type_               
            if type_ in ["row", "column", "cell"]:
                return column
            else:
                return None
        else:
            return None        
        
    def setSelectedCells(self, r1, r2, c1, c2):
        try:
            self.tksheet.set_currently_selected(r1, c1)
            self.tksheet.create_selection_box(r1, c1, r2, c2)
        except BaseException as e:
            logging.debug(e, exc_info=True) 
            logging.debug("setSelectedCells %s,%s,%s,%s", r1, r2, c1, c2)            
    
    def setSelectedShapes(self,shape):
        self.selectedShapes=[shape]
        
    def getSelectedShapes(self):
        return self.selectedShapes
    
    def is_column_hidden(self, col):
        if ActiveSheet.tksheet.displayed_columns != []:
            return not col in ActiveSheet.tksheet.displayed_columns
        else:
            return False        
    
    def set_hiderow(self,row,newval):
        if newval==True:
            #if not row-1 in ActiveSheet.tksheet.hiderowslist:
            if not ActiveSheet.is_row_hidden(row-1):
                #ActiveSheet.tksheet.hiderowslist.append(row-1)
                #ActiveSheet.Redraw_table()
                ActiveSheet.tksheet.hide_rows(rows=row-1)
        else:
            #if row-1 in ActiveSheet.tksheet.hiderowslist:
            if ActiveSheet.is_row_hidden(row-1):
                
                displayed_rows = ActiveSheet.tksheet.displayed_rows
                displayed_rows.append(row-1)
                displayed_rows.sort()
                ActiveSheet.tksheet.display_rows(displayed_rows, all_displayed=False, redraw=True)
                #ActiveSheet.tksheet.hiderowslist.remove(row-1)
                ActiveSheet.Redraw_table()
                
    def set_hidecolumn(self,col,newval):
        if newval==True:
            #if not col-1 in ActiveSheet.tksheet.hidecolslist:
                #ActiveSheet.tksheet.hidecolslist.append(col-1)
            if not ActiveSheet.is_column_hidden(col-1):
                ActiveSheet.tksheet.hide_columns(col-1)                
                ActiveSheet.Redraw_table()
        else:
            #if col-1 in ActiveSheet.tksheet.hidecolslist:
                #ActiveSheet.tksheet.hidecolslist.remove(col-1)
            if ActiveSheet.is_column_hidden(col-1):
                ActiveSheet.tksheet.displayed_columns.append(col-1)                
                ActiveSheet.Redraw_table()        
        
    def addrow_if_needed(self,row=None):
        if row==None:
            row = self.getSelectedRow()
        if row==self.getRowCount():
            self.addrow_before_current_row()
        
    def addrow_before_current_row(self,copy=False):
        try:
            cur_row = self.getSelectedRow()
            rows = self.tksheet.get_selected_rows(get_cells_as_rows=True)
            copyfromrow=None
            if copy:
                copyfromrow = cur_row
                self.tksheet.insert_rows(rows=len(rows),idx=cur_row)
                # copy content from old rows to new rows self.table.copy_row(rows=num_rows,idx=cur_row, from_row=copyfromrow)
                copyxx()
            self.tksheet.insert_rows(rows=len(rows),idx=cur_row)
        except BaseException as e:
            logging.debug("deleterows:", e, exc_info=True)        
    
    def deleterows(self):
        try:
            cur_row = self.getSelectedRow()
            rows = self.tksheet.get_selected_rows(get_cells_as_rows=True, return_tuple=True)
            if len(rows) == 1:
                self.tksheet.del_rows(cur_row)
            else:
                self.tksheet.del_rows(rows)
            if len(rows) > 0:
                self.deleteShapeatRow(rows[0], rows[-1])
                scrow_x1, scrow_y1, scrow_x2, scrow_y2 = self.getCellCoords(rows[0], 0)
                scrow2_x1, scrow2_y1, scrow2_x2, scrow2_y2 = self.getCellCoords(rows[-1], 0)
                minY1 = scrow_y1
                maxY1 = scrow2_y2
                deltaY = scrow2_y2 - scrow_y1
                self.moveShapesVertical(minY1, y2=maxY1, deltaY=deltaY)
        except BaseException as e:
            logging.debug("deleterows:", e, exc_info=True)
        
    def int_moveRows(self, sc_rowlist, destrow):
        self.tksheet.move_rows(move_to=destrow, to_move=sc_rowlist, move_data=True, create_selections=False)
        destrow_x1, destrow_y1, destrow_x2, destrow_y2 = self.getCellCoords(destrow, 0)
        scrow_x1, scrow_y1, scrow_x2, scrow_y2 = self.getCellCoords(sc_rowlist[0], 0)
        scrow2_x1, scrow2_y1, scrow2_x2, scrow2_y2 = self.getCellCoords(sc_rowlist[-1], 0)
        minY1 = scrow_y1
        maxY1 = scrow2_y2
        deltaY = destrow_y1 - scrow_y1
        deleteY = scrow2_y2 - scrow_y1
        self.moveShapesVertical(minY1, y2=maxY1, deltaY=deltaY)
        #self.moveShapesVertical(minY1,deltaY=-deleteY)
        return
    
    def prepare_moveRows(self):
        self.lastselectedRows = self.tksheet.get_selected_rows(get_cells_as_rows=True, return_tuple=True)
        
    def do_moveRows(self):
        self.Flag_move=True
        return
    
    def UsedRange_Rows(self):
        rowlist=[]
        for row in range(1,self.get_LastUsedRow()+1):
            rowlist.append(CRow(row))
        return rowlist
        
    def get_LastUsedRow(self):
        self.LastUsedRow_val = self.tksheet.get_total_rows() #test getLastUsedRow()+1 #self.tksheet.getRowCount()
        return self.LastUsedRow_val
    
    def set_LastUsedRow(self,value):
        self.LastUsedRow_val = value
        self.tksheet.setLastUsedRow(value-1)
    
    LastUsedRow = property(get_LastUsedRow, set_LastUsedRow, doc='Last used row')
    
    def get_LastUsedColumn(self):
        self.LastUsedColumn_val = self.tksheet.get_total_columns() #getColumnCount()
        return self.LastUsedColumn_val
    
    def set_LastUsedColumn(self,value):
        self.LastUsedColumn_val= value
    
    LastUsedColumn = property(get_LastUsedColumn, set_LastUsedColumn, doc='Last used column')
    End = property(get_LastUsedColumn, set_LastUsedColumn, doc='Last used column')

    def get_UsedRange(self):
        self.UsedRange_val = CRange((0,0) , (self.get_LastUsedRow(),self.LastUsedColumn_val),ws=self)
        return self.UsedRange_val
    
    def set_UsedRange(self,value):
        self.UsedRange_val = value
               
    UsedRange = property(get_UsedRange, set_UsedRange, doc='used range')
    MaxRange  = property(get_UsedRange, set_UsedRange, doc='used range')
    
    def get_Shape(self,shapename):
        for Shape in self.Shapes:
            if Shape.Name == shapename:
                return Shape
        return None
    
    def deleteShapeatPos(self,x1,y1,x2,y2):
        deleteList=[]
        index=0
        for shape in self.shapelist:
            if shape.Left in range(int(x1),int(x2)) and shape.Top in range(int(y1),int(y2)):
                deleteList.insert(0, index)
        if len(deleteList)>0:
            for item in deleteList:
                del self.shapelist[item]
                
    def deleteShapeatRow(self,y1,y2):
        for shape in self.shapelist:
            if shape.TopLeftCell_Row in range(y1,y2):
                shape.set_activeflag(False)
        #        deleteList.insert(0, index)
        #if len(deleteList)>0:
        #    for item in deleteList:
        #        self.shapelist[item].Active=False
        #test self.moveShapesVertical(y1,deltaY=y1-y2)
        #test self.setDataChanged()
        
    def copyShape(self,shape,newY=None):
        newshape = copy.copy(shape)
        newshape.Top = newY
        self.shapelist.append(newshape)
        #test self.setDataChanged()
            
    def moveShapesVertical(self,y1=-1,y2=-1,top_left_cell_row=None, deltaY=0,copy=False,cY1=0):
        tmp_Shapelist = self.shapelist.copy()
        if top_left_cell_row != None:
            for shape in tmp_Shapelist:
                if shape.TopLeftCell_Row != None:
                    if shape.TopLeftCell_Row > top_left_cell_row:
                        shape.Top += deltaY
        else:
            for shape in tmp_Shapelist:
                if y2==-1:
                    if shape.Top >=y1:
                        if shape.Top == cY1+1 and copy:
                            self.copyShape(shape,newY=shape.Top+deltaY)
                        else:
                            shape.Top += deltaY
                else:
                    if shape.Top in range(y1,y2+1):
                        if copy:
                            self.copyShape(shape,newY=shape.Top+deltaY)
                        else:
                            shape.Top += deltaY
        self.drawShapes()
        
    def moveShapesHorizontal(self,x1=-1,x2=-1,top_left_cell_column=None, deltaX=0,copy=False,cX1=0):
        tmp_Shapelist = self.shapelist.copy()
        if top_left_cell_column != None:
            for shape in tmp_Shapelist:
                if shape.TopLeftCell_Col != None:
                    if shape.TopLeftCell_Col > top_left_cell_column:
                        shape.Left += deltaX
        else:
            for shape in tmp_Shapelist:
                if x2==-1:
                    if shape.Top >=x1:
                        if shape.Left == cX1+1 and copy:
                            self.copyShape(shape,newX=shape.Top+deltaX)
                        else:
                            shape.Left += deltaX
                else:
                    if shape.Left in range(x1,x2+1):
                        if copy:
                            self.copyShape(shape,newX=shape.Left+deltaX)
                        else:
                            shape.Left += deltaX
        self.drawShapes()        
    
    def moveshapes(self,delta_width,asofpos):
        # move all shapes by delta_width after asofpos
        for shape in self.shapelist:
            if shape.Left > asofpos:
                shape.Left -=delta_width    

    def findshape(self,name):
        for shape in self.shapelist:
            if shape.Name==name:
                return shape
        return None
    
    def addShape(self, name, shapetype, Left, Top, Width, Height, Fillcolor, Text="",controls_dict=None):
        newshape = self.findshape(name)
        if newshape:
            return newshape
        filltkcolor=int2tkcolor(Fillcolor)
        newshape = CShape(name, shapetype, Left, Top, Width, Height, Fillcolor=filltkcolor, Text=Text, ws=self)
        self.shapelist.append(newshape)
        return newshape
    
    def addLabel(self, textorientation, Left, Top, Width, Height):
        name="IntLabel1"
        shapetype = msoTextBox
        filltkcolor="#000000"
        newshape = CShape(name, shapetype, Left, Top, Width, Height, Fillcolor=filltkcolor, Text=None, ws=self)
        self.shapelist.append(newshape)
        return newshape
    
    def addConnector(self, contype, BeginX, Beginy, Width, Height):
        name="IntConnector1"
        shapetype = "Connector"
        filltkcolor="#000000"
        newshape = CShape(name, shapetype, BeginX, Beginy, Width, Height, Fillcolor=filltkcolor, Text=None, ws=self)
        self.shapelist.append(newshape)
        return newshape    
    
    def addPicture(self, name, Left, Top, Width, Height, Text=None):
        shapetype = msoPicture
        filltkcolor="#000000"
        newshape = CShape(name, shapetype, Left, Top, Width, Height, Fillcolor=filltkcolor, Text=Text, ws=self)
        self.shapelist.append(newshape)
        return newshape
    
    def deleteshape(self,shape):
        if type(shape)==int:
            shape=self.shapelist[shape-1]
            shape.set_activeflag(False)
            self.shapelist.remove(shape) 
        else:
            if shape.rectidx!=0:
                self.tksheet.MT.delete(shape.rectidx)
            if shape.formwin!=None:
                self.tksheet.MT.delete(shape.formwin)
            self.shapelist.remove(shape) 
            shape.set_activeflag(False)

    def drawShapes(self,force=False):
        for shape in self.shapelist:
            self.drawShape(shape,force=force)
            
    def drawShape(self,shape,force=False):
        if shape.get_activeflag() and shape.get_visible():
            shapeY = shape.Top
            shaperow = shape.TopLeftCell_Row
            #determine displayed row index
            displayed_rows = self.tksheet.displayed_rows
            if displayed_rows == []:
                shaperow_displayed = True
            else:
                try:
                    shaperow = displayed_rows.index(shaperow)
                    shaperow_displayed = True
                except:
                    shaperow_displayed = False
                    #print("shape not in displayed row:", shaperow)
            if shaperow_displayed:
                if shape.rectidx==0 or force:
                    if shape.Shapetype == msoShapeRectangle or shape.Shapetype == "rect":
                        if shape.rectidx != 0:
                            self.tksheet.MT.delete(shape.rectidx)
                            pass # delete shape
                        #if shape.TextFrame2.TextRange.Text == "":
                        #    shape.TextFrame2.TextRange.Text = "0"
                        shape.rectidx = self.tksheet.MT.create_rectangle(shape.Left,shapeY,shape.Left+shape.Width,shapeY+shape.Height,fill=shape.Fillcolor,tags=(shape.Name,"Shape","Shapelist"))
                        if shape.ZOrder_Val==0:
                            self.tksheet.MT.tag_raise(shape.rectidx)
                        else:
                            self.tksheet.MT.tag_lower(shape.rectidx)
                        self.tksheet.MT.tag_bind(shape.rectidx,"<Button-1>", shape.shape_button_1)
                        if shape.Text !="":
                            if shape.textidx != 0:
                                self.tksheet.MT.delete(shape.textidx)
                            shape.textidx = self.tksheet.MT.create_text(int(shape.Left+shape.Width/2),int(shapeY+shape.Height/2),width=shape.Left+shape.Width,text=shape.Text,font=self.thefont,tags=(shape.Name,"Shape","Shapelist"))
                            if shape.ZOrder_Val==0:
                                self.tksheet.MT.tag_raise(shape.textidx)
                            else:
                                self.tksheet.MT.tag_lower(shape.textidx)                        
                            self.tksheet.MT.tag_bind(shape.textidx,"<Button-1>", shape.shape_button_1)
                        else:
                            shape.textidx = -1
                    elif shape.Shapetype == msoTextBox or shape.Shapetype == "TextBox":
                        #if shape.TextFrame2.TextRange.Text == "":
                        #    shape.TextFrame2.TextRange.Text = "0"
                        if shape.Text !="":
                            shape.textidx = self.tksheet.MT.create_text(int(shape.Left),int(shape.Top),anchor="nw",text=shape.Text,tags=(shape.Name,"Shape","Shapelist"),width=shape.Width,font=self.thefont)
                            if shape.ZOrder_Val==0:
                                self.tksheet.MT.tag_raise(shape.textidx)
                            else:
                                self.tksheet.MT.tag_lower(shape.textidx)                        
                            #self.tag_bind(shape.textidx,"<Button-1>", shape.shape_button_1)
                        else:
                            shape.textidx = -1
                    elif shape.Shapetype == msoOLEControlObject:
                        if shape.formwin==None or force:
                            formFrame = ttk.Frame(self.tksheet,borderwidth=0)
                            control_dict = shape.control_dict.get("Components",None)
                            format_dict  = shape.format_dict
                            shape.AlternativeText = shape.control_dict.get("AlternativeText",None)
                            generate_controls(control_dict,formFrame,self.worksheet,persistent_controls=self.persistent_controls_dict,format_dict=format_dict,defaultfont=self.thefont, jump_table=self.Workbook.jumptable)
                            if shape.Top==None:
                                x1,y1,x2,y2 = self.getCellCoords(shape.Row-1,shape.Col-1)
                                shape.Top=y1
                                shape.Left=x1
                            else:
                                x1 = shape.Left
                                y1 = shape.Top
                            #shape.form.bind('<Return>', callback)
                            #shape.form.bind('<KeyRelease>', callback)
                            #shape.form.bind("<Button-3>", self.handle_right_click)
                            #self.form.focus_set()
                            shape.formwin=self.tksheet.MT.create_window(x1,y1, width=shape.Width,height=shape.Height,window=formFrame,anchor='nw',tag=('form',"Shapelist"))
                        else:
                            self.tksheet.MT.coords(shape.formwin,shape.Left, shape.Top)
                            if shape.updatecontrol:
                                control = self.worksheet.Controls_Dict.get(shape.Name,None)
                                if control:
                                    control.TKWidget.configure(text=shape.Text,background=shape.Fillcolor)
                                    #control.TKWidget.update()
                        self.tksheet.MT.tag_raise(shape.formwin)
                    elif shape.Shapetype == "picture" or shape.Shapetype==msoPicture:
                        pass
                        image_loaded=False
                        try:
                            image = tk.PhotoImage(file=shape.Name)
                            self.imagedict[shape.Name+str(shape.Top)] = image
                            shape.rectidx = self.tksheet.MT.create_image(shape.Left,shape.Top,image=image,tags=(shape.Name,"Shape","Picture","Shapelist"),anchor='nw')
                            self.tksheet.MT.tag_raise(shape.rectidx)
                            image_loaded=True
                        except BaseException as e:
                            #logging.debug(e, exc_info=True) 
                            #logging.debug("Worksheet: Error loading image:" + shape.Name)
                            image_loaded=False
                        if not image_loaded:
                            try:
                                shape.Name=shape.Name.replace(".jpg",".png")
                                image = tk.PhotoImage(file=shape.Name)
                                self.imagedict[shape.Name+str(shape.Top)] = image
                                shape.rectidx = self.tksheet.MT.create_image(shape.Left,shape.Top,image=image,tags=(shape.Name,"Shape","Picture","Shapelist"),anchor='nw')
                                self.tksheet.MT.tag_raise(shape.rectidx)
                                image_loaded=True
                            except BaseException as e:
                                #logging.debug(e, exc_info=True) 
                                #logging.debug("Worksheet: Error: Image not found: "+shape.Name)
                                image_loaded = False
                        if not image_loaded:
                            try:
                                shape.Name=shape.Name.replace(".bmp",".png")
                                image = tk.PhotoImage(file=shape.Name)
                                self.imagedict[shape.Name+str(shape.Top)] = image
                                shape.rectidx = self.tksheet.MT.create_image(shape.Left,shape.Top,image=image,tags=(shape.Name,"Shape","Picture","Shapelist"),anchor='nw')
                                self.tksheet.MT.tag_raise(shape.rectidx)
                                image_loaded=True
                            except BaseException as e:
                                logging.debug(e, exc_info=True) 
                                logging.debug("Worksheet: Error: Image not found: "+shape.Name)
                                image_loaded = False
                        #self.tag_raise(shape.textidx)
                        #self.tag_bind(shape.rectidx,"<Button-1>", shape.shape_button_1)
                        #self.tag_bind(shape.textidx,"<Button-1>", shape.shape_button_1)
                    elif shape.Shapetype == msoShapeOval or shape.Shapetype == "oval":
                        if shape.TextFrame2 == "":
                            shape.TextFrame2 = "0"
                        shape.rectidx = self.tksheet.MT.create_oval(shape.Left,shapeY,shape.Left+shape.Width,shapeY+shape.Height,fill=shape.Fillcolor,tags=(shape.Name,"Shape","Shapelist"))
                        self.tksheet.MT.tag_raise(shape.rectidx)
                        if shape.Text !="":
                            shape.textidx = self.tksheet.MT.create_text(int(shape.Left+shape.Width/2),int(shapeY+shape.Height/2),width=shape.Left+shape.Width,text=shape.Text,font=self.thefont,tags=(shape.Name,"Shape","Shapelist"))
                            self.tksheet.MT.tag_raise(shape.textidx)
                        #self.tag_bind(shape.rectidx,"<Button-1>", shape.shape_button_1)
                        #self.tag_bind(shape.textidx,"<Button-1>", shape.shape_button_1)
                    elif shape.Shapetype == msoFreeform:
                        if shape.SegmentType == msoSegmentCurve:
                            smooth=True
                        else:
                            smooth=False
                        if shape.EditingType == msoEditingAuto:
                            shape.rectidx = self.tksheet.MT.create_polygon(shape.nodelist,fill=shape.Fillcolor,outline=shape.LineColor,smooth=smooth,tags=(shape.Name,"Shape","Shapelist"))
                        else:
                            arrow=None
                            if hasattr(shape,"EndArrowheadStyle_val"):
                                if shape.EndArrowheadStyle_val in (msoArrowheadStealth,msoArrowheadTriangle):
                                    arrow=tk.LAST
                            if hasattr(shape,"BeginArrowheadStyle_val"):
                                if shape.BeginArrowheadStyle_val in (msoArrowheadStealth,msoArrowheadTriangle):
                                    arrow=tk.FIRST                               
                            shape.rectidx = self.tksheet.MT.create_line(shape.nodelist,fill=shape.Fillcolor,smooth=smooth,arrow=arrow,tags=(shape.Name,"Shape","Shapelist"))
                        if True: #shape.ZOrder_Val==0:
                            self.tksheet.MT.tag_raise(shape.rectidx)
                        else:
                            self.tksheet.MT.tag_lower(shape.rectidx)
                        if shape.Text !="":
                            shape.textidx = self.tksheet.MT.create_text(int(shape.Left+shape.Width/2),int(shapeY+shape.Height/2),width=shape.Left+shape.Width,text=shape.Text,tags=(shape.Name,"Shape","Shapelist"))
                            self.tksheet.MT.tag_raise(shape.textidx)                    
                    elif shape.Shapetype == "Connector":
                        #if shape.TextFrame2.TextRange.Text == "":
                        #    shape.TextFrame2.TextRange.Text = "0"
                        arrow=None
                        if hasattr(shape,"EndArrowheadStyle_val"):
                            if shape.EndArrowheadStyle_val == msoArrowheadStealth:
                                arrow=tk.LAST
                        if hasattr(shape,"BeginArrowheadStyle_val"):
                            if shape.BeginArrowheadStyle_val == msoArrowheadStealth:
                                arrow=tk.FIRST                        
                        shape.rectidx = self.tksheet.MT.create_line(shape.Left,shape.Top,shape.Left+shape.Width,shape.Top+shape.Height,fill=shape.Fillcolor,arrow=arrow,tags=(shape.Name,"Shape","Shapelist"))
                        if shape.ZOrder_Val==0:
                            self.tksheet.MT.tag_raise(shape.rectidx)
                        else:
                            self.tksheet.MT.tag_lower(shape.rectidx)
                    else:
                        logging.debug("Unknown Shapetype: %s",shape.Shapetype)
                else:
                    if shape.ZOrder_Val==0:
                        self.tksheet.MT.tag_raise(shape.rectidx)
                    else:
                        self.tksheet.MT.tag_lower(shape.rectidx)
                    if shape.Width==None or shape.Height==None:
                        params = self.tksheet.MT.coords(shape.rectidx)
                        self.tksheet.MT.coords(shape.rectidx,shape.Left, shape.Top)
                    else:
                        params = self.tksheet.MT.coords(shape.rectidx)
                        if len(params) == 2:
                            self.tksheet.MT.coords(shape.rectidx,shape.Left, shape.Top)
                        else:
                            self.tksheet.MT.coords(shape.rectidx,shape.Left, shape.Top,shape.Left+shape.Width,shape.Top+shape.Height)
                    if shape.textidx!=0:
                        if shape.ZOrder_Val==0:
                            self.tksheet.MT.tag_raise(shape.textidx)
                        else:
                            self.tksheet.MT.tag_lower(shape.textidx)
                        self.tksheet.MT.itemconfigure(shape.textidx, text=shape.Text)
                        self.tksheet.MT.coords(shape.textidx,int(shape.Left+shape.Width/2),int(shape.Top+shape.Height/2))
    
    def drawGrid(self, startrow=None, endrow=None):
        #Draw the table grid lines
        self.tksheet.MT.delete('gridline')
        if self.Gridlist != {}:
            #print("Grid_ Startrow - Endrow:",startrow,endrow)
            for grid in self.Gridlist:
                griditem = self.Gridlist[grid]
                cell0 = griditem[0]
                cell1 = griditem[1]
                row0 = cell0[0] - 1
                col0 = cell0[1] - 1
                row1=cell1[0] - 1
                col1=cell1[1] - 1
                startrow1=self.data_row_to_displayed(row0)
                endrow1=self.data_row_to_displayed(row1)
                if startrow1 != None and endrow1 != None:
                    y_start = self.calc_y_for_row(startrow1)
                    y_end = self.calc_y_for_row(endrow1+1)
                    for col in range(col0,col1+2):
                        x=self.calc_x_for_column(col)
                        self.tksheet.MT.create_line(x,y_start,x,y_end, tag='gridline',
                                             fill=self.grid_color, width=self.linewidth)
                    
                    x_start = self.calc_x_for_column(col0)
                    x_end   = self.calc_x_for_column(col1+1)
                    y_pos = y_start
                    for row in range(startrow1, endrow1+2):
                        #print("Grid Horizontal - y_pos:",row,y_pos)
                        y_pos =  self.calc_y_for_row(row)
                        self.tksheet.MT.create_line(x_start,y_pos,x_end,y_pos, tag='gridline',
                                        fill=self.grid_color, width=self.linewidth)
                        
    
    def getShape(self,index):
        if index <= len(self.shapelist):
            shape = self.shapelist[index-1]
        else:
            shape=None
        return shape
    
    def setDataChanged(self):
        pass
    
    def data_row_to_displayed(self, row):
        try:
            displayed_row = self.tksheet.displayed_rows.index(row)
        except:
            displayed_row = None
        return displayed_row
    
    def displayed_row_to_data(self, row):
        return self.tksheet.displayed_row_to_data(row)
    
    def data_column_to_displayed(self, col):
        displayed_col = self.tksheet.displayed_columns.index(col)
        return displayed_col
    
    def displayed_column_to_data(self, col):
        return self.tksheet.displayed_column_to_data(col)
    
    def importCSV(self, filename=None, sep=',',fieldnames=None):
        #Import from csv file
        self.importCSV_int(filename, sep=sep,fieldnames=fieldnames)
        self.shapelist=[]
        #test self.updateModel(model)
        return
    
    def importCSV_int(self, filename, sep=',',fieldnames=None):
        #Import table data from a comma separated file.
        
        logging.debug("Import CSV_int: Filename="+str(filename))

        if not os.path.isfile(filename) or not os.path.exists(filename):
            #print ('no such file', filename)
            logging.debug("Import CSV_int: File notfound Filename="+filename)
            return None
        #takes first row as field names
        dictreader = csv.DictReader(open(filename, "r",encoding="utf8"), delimiter=sep,fieldnames=fieldnames)
        dictdata = []
        for rec in dictreader:
            recdata = []
            for cell, value in rec.items():
                recdata.append(value)
            dictdata.append(recdata)
        self.importDict(dictdata)
        self.modelname=filename
        return

    def importDict(self, newdata):
        #test needs to be updated
        #get cols from sub data keys
        #logging.debug("importDICT:"+repr(newdata))
        self.tksheet.data = newdata
        self.setDataChanged()
        return
    
    def refreshicons(self):
        
        for row in range(0, 10):
            s = self.Cells(row, 12)
            if s != '':
                M27.FindMacro_and_Add_Icon_and_Name(s, row, ActiveSheet)        
        M20.Update_Start_LedNr()
        
        self.Workbook.Evt_Redraw_Sheet(self)
        
    def clear_shapelist(self):
        self.tksheet.MT.delete("Shapelist")
        self.shapelist = []
    
    def recreate_controls(self):
        sheetname_prop = self.Workbook.sheetdict.get(self.Name)
        controls_dict = sheetname_prop.get("Controls",None)
        if controls_dict:
            self.add_controls(controls_dict)
           
    def Cells(self,row=None, col=None, ws=None):
        if ws == None:
            ws = self
        if row==None and col==None:
            return self.get_UsedRange()
        return CRange(row,col,ws=ws)
    
    def Rows(self,row):
        try:
            return self.wsRowdict[row]
        except:
            self.wsRowdict[row] = CRow(row)
            return self.wsRowdict[row]
    
    def get_cellvalue_direct(self,sheet,row,col):
        tksheet = sheet.tksheet
        max_cols = tksheet.get_total_columns() #getColumnCount()
        if col <= max_cols:
            value = tksheet.get_data(row-1,col-1)
        else:
            value = ""
        return value
    
    def find_RangeName(self,rangestr,ws=None):
        # returns CRange with row and column of the Named_range
        if ws==None:
            ws  = ActiveSheet
        #check for Cell Names eg. "E5", "AB12"
        pattern=re.compile(r"([A-Z]+)(\d+)")
        m = re.match(pattern,rangestr)
        if m!=None:
            #print(rangestr,"->",m.groups())
            m_list = m.groups(0)
            col_str = m_list[0]
            named_range_col = 0
            for char in col_str:
                index = ord(char)-ord("A")+1
                named_range_col = named_range_col*26+index
            row_str = m_list[1]
            named_range_row = int(row_str)
            resultcell = CRange(self.Cells(named_range_row,named_range_col, ws=self))
            return resultcell
        
        LSh = ws.Workbook.Sheets("Named_Ranges")
        _row = LSh.find_in_col_ret_row(rangestr,1,cache=True)        
        if _row != None:
            named_range_sheetname= self.get_cellvalue_direct(LSh,_row,2)
            named_range_sheet = ws.Workbook.Sheets(named_range_sheetname)
            named_range_col = int(self.get_cellvalue_direct(LSh,_row,3))
            named_range_row = int(self.get_cellvalue_direct(LSh,_row,4))
            value = self.get_cellvalue_direct(ws, named_range_row,named_range_col) 
            #logging.debug("FindRangeName: "+rangestr+" - "+ named_range_sheetname+"("+str(named_range_row)+","+str(named_range_col)+")"+"="+str(value))
            #logging.debug("FindRangeName-Current ws:"+ws.Name)
            resultcell = CRange(self.Cells(named_range_row,named_range_col, ws=self),value=value, ws=self)
            return resultcell
        else:
            return None

    def Range(self,cell1,cell2=None,ws=None):
        if ws==None:
            ws=self        
        return CRange(cell1,cell2,ws=ws)
        
    def Range_set(self,rangestr,value):
        oldflag = Application.EnableEvents
        Application.EnableEvents=False
        named_cell = self.find_RangeName(rangestr, ws=self)
        if named_cell == None: #test
            return #test
        named_cell.Value = value
        Application.EnableEvents = oldflag
    
    def Unprotect(self,Password=""):
        self.ProtectContents = False
        
    def Protect(self, DrawingObjects=True, Contents=True, Scenarios=True, AllowFormattingCells=True, AllowInsertingHyperlinks=True):
        self.ProtectContents = True
        return
        
    def Columns(col:int):
        #print("Columns",col)
        return CColumn(col)
    
    def EnableDisableAllButtons(self,value):
        return
    
    def SetChangedCallback(self,callback):
        self.wschanged_callback = callback
        
    def SetSelectedCallback(self,callback):
        self.wsselected_callback = callback
        
    def SetCalculationCallback(self,callback):
        self.wscalculation_callback = callback
        
    def SetInitDataFunction(self,func):
        self.wsInitDataFunction = func
        
    def EventWScalculate(self,changecell=None):
        if self.wscalculation_callback:
            self.wscalculation_callback(self,changecell)
       
    def EventWSchanged(self, changedcell):
        if self.wschanged_callback and Application.EnableEvents:
            if Application.EnableCalculation and self.EnableCalculation and Application.Calculation == xlCalculationAutomatic:
                Application.EnableEvents=False
                self.wsselected_callback(changedcell)
                self.wscalculation_callback(self,cell=changedcell) # Calc_Worksheet(self)
                self.wschanged_callback(changedcell)
                Application.EnableEvents=True
                self.Redraw_table()
            
    def EventWSselected(self, selectedcell):
        if self.wsselected_callback and Application.EnableEvents:
            self.wsselected_callback(selectedcell)
            
    def EventWSselected_orig(self, selectedcell):
        if self.wsselected_callback and Application.EnableEvents:
            self.wsselected_callback(selectedcell)    
            
    def EventWSdoubleclick(self, event):
        row,col = self.getSelectedCell()
        target = self.Cells(row, col)
        cancel = None
        self.Evt_SheetBeforeDoubleClick(self,target, cancel)
        
    def EventWSLeftClick(self, event):
        print("EventWSLeftClick:", event)
            
    def Redraw_table(self,do_bindings=False):
        #test needs update
        self.tksheet.redraw()
        self.drawShapes()
        self.drawGrid()
        if do_bindings:
            #print("Redraw_Table - do bindings")
            self.do_bindings()
        #self.Call_EventProc("Worksheet_Redraw")(self) 
        
    def set_value_in_cell(self,row,column,newval):
        cell = self.Cells(row, column)
        cell.Value=newval
        
    def create_search_colcache(self,searchcol):
        colcontent = self.tksheet.get_data(self.tksheet.span(None, (searchcol-1)))
        row = 1
        self.searchcache[searchcol]={}
        searchcol_dict = self.searchcache[searchcol]
        for col in colcontent:
            if len(col)>50:
                col=col[:50]
            searchcol_dict[col]=row
            row=row+1
        return
        
    def find_in_col_ret_col_val(self,searchtext, searchcol, resultcol,cache=True):
        if cache:
            colcache = self.searchcache.get(searchcol,None)
            if not colcache:
                self.create_search_colcache(searchcol)
                colcache = self.searchcache.get(searchcol,None)
            if len(searchtext)>50:
                searchtext=searchtext[:50]
            res_row = colcache.get(searchtext,None)
            if not res_row:
                return None # searchtext not found
            else:
                return self.getCellRecord(res_row-1, resultcol-1)
        else:
            pass
        return None
    
    def getCellRecord(self, row, col):
        value = self.tksheet.get_data(self.tksheet.span(row, col))
        return value
    
    def getCellCoords(self, r,c):
        try:
            new_r = self.tksheet.displayed_rows.index(r)
        except:
            new_r = r
        x1 = self.tksheet.MT.col_positions[c]
        y1 = self.tksheet.MT.row_positions[new_r]
        x2 = self.tksheet.MT.col_positions[c + 1] - 1
        y2 = self.tksheet.MT.row_positions[new_r + 1] - 1        
        return x1,y1,x2,y2
    
    def calc_row_from_y(self, y):
        displayed_row = self.tksheet.MT.identify_row(y=y)
        row = self.displayed_row_to_data(displayed_row)
        return row
    
    def calc_col_from_x(self, x):
        col = self.tksheet.MT.identify_col(x=x)
        return col
    
    def calc_y_for_row(self, row):
        row_pos_list = self.tksheet.get_row_heights(canvas_positions=True)
        y = row_pos_list[row]
        return y
    
    def calc_x_for_column(self, column):
        column_pos_list = self.tksheet.get_column_widths(canvas_positions=True)
        x = column_pos_list[column]
        return x    
    
    def find_in_col_ret_row(self,searchtext, searchcol, cache=True):
        if cache:
            colcache = self.searchcache.get(searchcol,None)
            if not colcache:
                self.create_search_colcache(searchcol)
                colcache = self.searchcache.get(searchcol,None)
            res_row = colcache.get(searchtext,None)
            if res_row == None:
                res_row = colcache.get(searchtext+" ",None)
            if not res_row:
                return None # searchtext not found
            else:
                return res_row
        return None    
                
    def find_in_col_set_col_val(self,searchtext, searchcol, setcol,setval,cache=False):
        if cache:
            colcache = self.searchcache.get(searchcol,None)
            if not colcache:
                self.create_search_colcache(searchcol)
                colcache = self.searchcache.get(searchcol,None)
            res_row = colcache.get(searchtext,None)
            if not res_row:
                return None # searchtext not found
            else:
                self.set_value_in_cell(res_row, setcol, setval)
                return True
        else:
            pass
        return None
    
    def Activate(self):
        global ActiveSheet,ActiveWorkbook
        #print("Worksheet.Activate",self.Workbook.Name, self.Name)
        set_activeworkbook(self.Workbook)
        ActiveSheet = self
        ActiveWorkbook = self.Workbook
        self.Workbook.ActiveSheet=self
        #self.Workbook.ActiveSheet =self
        #self.do_bindings()
        #if self.Workbook.Name=="ProgGenerator":
        #    Port=Cells(M02.SH_VARS_ROW, M25.COMPort_COL)
            #if F00.port_is_available(Port):    
            #    self.controller.setConfigData("serportname",Port)
        return

    def Copy(self,SheetName=None,After=None):
        self.Workbook.add_sheet(SheetName,from_sheet=self.Name,After=After)
        self.Workbook.activate_sheet(SheetName)
    
    def do_bindings(self):
        #print("do bindings")
        #test self.table.do_bindings()
        pass
        
    def remove_bindings(self):
        #print("remove bindings")
        #test self.table.remove_bindings()
        pass
    
    def Select(self):
        global ActiveSheet,Selection
        ActiveSheet = self
        #print("Worksheet.Select",self.Workbook.Name, self.Name)
        self.Workbook.ActiveSheet = self
        Selection=self.Workbook.Selection
        self.Workbook.container.select(self.tabid)
        return
    
    #def Visible(self,value):
    #    if value==True:
    #        self.Workbook.container.tab(self.tabid,state="normal")
    #        #self.Workbook.container.add(self)
    #    else:
    #        self.Workbook.container.tab(self.tabid,state="disable")
    #        self.Workbook.container.hide(self.tabid)
    #    return
    
    def get_visible(self):
        return self._visible_val

    def set_visible(self, value):
        self._visible_val=value
        if value==True:
            self.Workbook.container.tab(self.tabid,state="normal")
        else:
            self.Workbook.container.tab(self.tabid,state="disable")
            self.Workbook.container.hide(self.tabid)            

    Visible = property(get_visible, set_visible, doc='Visible Status')    

    def getData(self):
        #collects all table data and status information to allow the creation of a JSON file
        data = {}
        tksheet_data = self.tksheet.get_data(self.tksheet.span(""))
        
        data["TKSHEET_data"] = copy.deepcopy(tksheet_data)
        highlighted_cells = self.tksheet.get_highlighted_cells()
        if highlighted_cells != {}:
            new_highlighted_cells = {}
            for cell_tuple,highlight in highlighted_cells.items():
                new_key = str(cell_tuple[0]) + "," + str(cell_tuple[1])
                new_highlighted_cells[new_key] = highlight.bg
            data["TKSHEET_highlighted_cells"] = copy.deepcopy(new_highlighted_cells)
        data["row_pos"] = copy.deepcopy(self.tksheet.get_row_heights(canvas_positions=True))
        data["col_pos"] = copy.deepcopy(self.tksheet.get_column_widths(canvas_positions=True))
        if self.saveShapes:
            shapelist = self.shapelist
            new_shapelist = []
            for shape in shapelist:
                newshape = dict()
                newshape["WorksheetName"] = shape.Worksheet.Name
                newshape["Shapetype"] = shape.Shapetype
                newshape["Type"] = shape.Type
                newshape["Name_val"] = shape.Name_val
                newshape["tablename"] = shape.tablename
                newshape["Left_val"] = shape.Left_val
                newshape["Active"] = shape.Active
                newshape["Top_val"] = shape.Top_val
                newshape["Width"] = shape.Width
                newshape["Height"] = shape.Height
                newshape["Row"] = shape.Row
                newshape["Col"] = shape.Col
                newshape["Rotation"] = shape.Rotation
                newshape["Visible_val"] = shape.Visible_val
                newshape["TopLeftCell_Row"] = shape.TopLeftCell_Row
                newshape["TopLeftCell_Col"] = shape.TopLeftCell_Col
                newshape["Index"] = shape.Index
                newshape["OnAction_val"] = shape.OnAction_val
                newshape["ZOrder_Val"] = shape.ZOrder_Val
                newshape["Locked"] = shape.Locked
                newshape["Text"] = shape.Text
                newshape["AlternativeText"] = shape.AlternativeText
                #newshape["Fill.ForeColor"] = shape.Fill.ForeColor
                #newshape["Line.BackColor"] = shape.Line.BackColor
                #newshape["Line.ForeColor"] = shape.Line.ForeColor
                #newshape["Line.BeginArrowheadStyle_val"] = shape.Line.BeginArrowheadStyle_val
                #newshape["Line.EndArrowheadStyle_val"] = shape.Line.EndArrowheadStyle_val
                #newshape["Line.Transparency"] = shape.Line.Transparency
                #newshape["Line.Visible"] = shape.Line.Visible
                #newshape["Line.Weight"] = shape.Line.Weight
                newshape["control_dict"] = shape.control_dict
                newshape["updatecontrol"] = shape.updatecontrol
                newshape["AlternativeText"] = shape.AlternativeText
                newshape["Init_Value"] = shape.Init_Value
                newshape["Fillcolor"] = shape.Fillcolor
                newshape["LineColor"] = shape.LineColor
                newshape["nodelist"] = shape.nodelist
                newshape["SegmentType"] = shape.SegmentType
                newshape["EditingType"] = shape.EditingType
                new_shapelist.append(newshape)
            data["shapelist"] = copy.deepcopy(new_shapelist)
            
        self.SavePersistentControls()
        data["Persistent_Controls"]=copy.deepcopy(self.persistent_controls_dict)
        return data
    
    def setData(self,data):
        sheet_data = copy.deepcopy(data.get("TKSHEET_data", {}))
        #self.table.set_data(data["TKSHEET_data"])
        self.tksheet.set_data("A1", data=sheet_data, emit_event=True, redraw=True)
        
        highlighted_cells = data.get("TKSHEET_highlighted_cells", {})
        if highlighted_cells != {}:
            for cell_str,highlight in highlighted_cells.items():
                cell_split = cell_str.split(",")
                new_key = (int(cell_split[0]), int(cell_split[1]))
                self.tksheet.highlight(new_key, bg=highlight)

        shapelist = copy.deepcopy(data.get("shapelist", []))
        if shapelist != []:
            new_shapelist = []
            for newshape in shapelist:
                shape = CShape(newshape["Name_val"], newshape["Shapetype"], newshape["Left_val"], newshape["Top_val"], Width=0, Height=0, Fillcolor=0, Text="",Row=None, Col=None, rotation=0,orientation=msoTextOrientationHorizontal,AlternativeText="",ws=None,nodelist=None,control_dict=None,Init_Value=None,linecolor=None,zorder=0,segmenttype=msoSegmentLine,editingtype=msoEditingAuto)
                shape.Worksheet = self
                shape.Shapetype = newshape["Shapetype"]
                shape.Type = newshape["Type"]
                shape.Name_val = newshape["Name_val"]
                shape.tablename = newshape["tablename"]
                shape.Left_val = newshape["Left_val"]
                shape.Active = newshape["Active"]
                shape.Top_val = newshape["Top_val"]
                shape.Width = newshape["Width"]
                shape.Height = newshape["Height"]
                shape.Row = newshape["Row"]
                shape.Col = newshape["Col"]
                shape.Rotation = newshape["Rotation"]
                shape.Visible_val = newshape["Visible_val"]
                shape.TopLeftCell_Row = newshape["TopLeftCell_Row"]
                shape.TopLeftCell_Col = newshape["TopLeftCell_Col"]
                shape.Index = newshape["Index"]
                shape.OnAction_val = newshape["OnAction_val"]
                shape.ZOrder_Val = newshape["ZOrder_Val"]
                shape.Locked = newshape["Locked"]
                shape.Text = newshape["Text"]
                shape.AlternativeText = newshape["AlternativeText"]
                shape.Fill = CFill(newshape["Fillcolor"], shape=shape)
                shape.Line = CLine(linecolor= newshape["LineColor"], shape=self)
                #shape.Line.BackColor = newshape["Line.BackColor"]
                #shape.Line.ForeColor = newshape["Line.ForeColor"]
                #shape.Line.BeginArrowheadStyle_val = newshape["Line.BeginArrowheadStyle_val"]
                #shape.Line.EndArrowheadStyle_val = newshape["Line.EndArrowheadStyle_val"]
                #shape.Line.Transparency = newshape["Line.Transparency"]
                #shape.Line.Visible = newshape["Line.Visible"]
                #shape.Line.Weight = newshape["Line.Weight"]
                shape.control_dict = newshape["control_dict"]
                shape.updatecontrol = newshape["updatecontrol"]
                shape.AlternativeText = newshape["AlternativeText"]
                shape.Init_Value = newshape["Init_Value"]
                shape.Fillcolor = newshape["Fillcolor"]
                shape.LineColor = newshape["LineColor"]
                shape.nodelist = newshape["nodelist"]
                shape.SegmentType = newshape["SegmentType"]
                shape.EditingType = newshape["EditingType"]
                new_shapelist.append(shape)
            self.shapelist = new_shapelist

        row_pos = data.get("row_pos", [])
        if row_pos != []:
            self.tksheet.set_row_heights(row_heights=row_pos, canvas_positions=True)
            
        col_pos = data.get("col_pos", [])
        if col_pos != []:
            self.tksheet.set_column_widths(column_widths=col_pos, canvas_positions=True)

        self.persistent_controls_dict=data.get("Persistent_Controls",{})
        self.SetPersistentControls()
        
        self.tksheet.redraw() #Table(force=True)
        self.drawShapes(force=True)
        self.drawGrid()

    def clearData(self):
        self.tksheet.reset()
        
        
    def importdefaultfile(self):
        if self.csv_filepathname != None:
            fileextension = os.path.splitext(self.csv_filepathname)[1]
            fileextension = fileextension.upper()
            if fileextension == ".CSV":
                self.importCSV(filename=self.csv_filepathname, sep=';',fieldnames=self.fieldnames)
            elif fileextension == ".MLL_PCF":
                self.Workbook.load()
            elif fileextension == ".XLSM":
                self.Workbook.LoadSheetfromExcelWorkbook(self,self.csv_filepathname, self.fieldnames)            
    
    def clearSheet(self):
        #test needs to be updated
        self.clearData()
        self.importdefaultfile()
        #test self.tablemodel = self.table.getModel()
        if self.formating_dict:
            #self.tablemodel.nodisplay = self.formating_dict.get("HideCells",[])
            #self.tablemodel.hiderowslist = self.formating_dict.get("HideRows",[])
            #self.tablemodel.protected_cells = self.formating_dict.get("ProtectedCells",[])
            #self.tablemodel.format_cells = self.formating_dict.get("FontColor",{})
            #test self.tablemodel.nodisplay = self.formating_dict.get("HideCells",[])
            self.handle_hidden_cells(self.formating_dict.get("HideCells",[]))
            #test self.tablemodel.hiderowslist = self.formating_dict.get("HideRows",[])
            self.handle_hidden_rows(self.formating_dict.get("HideRows",[]))
            #test self.tablemodel.protected_cells = self.formating_dict.get("ProtectedCells",[])
            self.handle_protected_cells(self.formating_dict.get("ProtectedCells",[]))
            #test self.tablemodel.format_cells = self.formating_dict.get("FontColor",{})
            self.handle_formating(self.formating_dict.get("FontColor",[]))
            #test self.tablemodel.gridlist = self.formating_dict.get("GridList",[])
            self.handle_gridlist(self.formating_dict.get("GridList",[]))
            self.handle_gridformat(self.formating_dict.get("GridFormat",{}))
            #test self.tablemodel.ColumnAlignment = self.formating_dict.get("ColumnAlignment",{})
            self.handle_ColumnAlignment(self.formating_dict.get("ColumnAlignment",[]))
            self.handle_ColumnWidth(self.formating_dict.get("ColumnWidth",[]))            
        #self.table.redraw()
        self.update_table_properties()
        self.Activate()
        self.controller.update()
        
    def check_Data_Changed(self):
        return self.DataChanged #test or self.tablemodel.checkDataChanged()
    
    def add_new_attr(self,attr):
        setattr(self,attr,attr)
        
    def add_UI_object(self,ui_object_name,UI_object):
        setattr(self,ui_object_name, UI_object)
        
    def changeGrid(self, gridno, diditem):
        #test needs update
        self.Gridlist[gridno] = diditem
        
    def get_selection(self):
        return CRange(self.getSelectedCell())
    def set_selection(self, value):
        ActiveSheet.get_selection().Value=value
    Selection = property(get_selection, set_selection, doc='Selection')
    
    def get_name(self):
        return self.name_val
    def set_name(self, value):
        self.name_val = value
    Name = property(get_name, set_name, doc='Name')
    
    def getColorAt(self, row, col):
        return self.tksheet.getColorAt(row, col)
 
 
class CFreeForm(object):
    def __init__(self, EditingType, X1, Y1):
        self.nodelist=[(X1,Y1)]
        self.EditingType=EditingType
        self.Left=X1
        self.Top=Y1
        self.SegmentType = None
        
    def AddNodes(self,SegmentType, EditingType, X2, Y2, *args):
        self.SegmentType = SegmentType
        self.EditingType = EditingType
        point=(X2,Y2)
        self.nodelist.append(point)
        get_xval=True
        for val in args:
            if get_xval:
                x=val
                get_xval=False
            else:
                y=val
                point=(x,y)
                self.nodelist.append(point)
                get_xval=True
        
    def ConvertToShape(self):
        name="Freeform"
        shape=CShape(name, msoFreeform, self.Left, self.Top, None, None,Fillcolor=-1,nodelist=self.nodelist,zorder=0,segmenttype=self.SegmentType,editingtype=self.EditingType)
        #print("ConvertToShape", name, self.Left, self.Top, self.nodelist)
        shape.Worksheet.shapelist.append(shape)
        #print(shape.Worksheet.shapelist)
        return shape
        
class CShapeList(object):
    def __init__(self,ws=None):
        if ws==None:
            ws=ActiveSheet
        self.ws = ws
        self.ws.shapelist = list()
        self.currentshape=None
        self.currentidx = 0

    def getShape(self,index):
        if type(index)==int:
            if index <= len(self.ws.shapelist):
                shape = self.ws.shapelist[index-1]
            else:
                shape=None
        else:
            shape=self.ws.get_Shape(index) # by name
        return shape
    
    def getlist(self):
        return self.ws.shapelist
        
    def AddShape(self, shapetype, Left, Top, Width, Height, Row=None, Col=None, Fill=0,text="",name="",control_dict=None,Init_Value=None):
        shape=None
        if shapetype == msoShapeRectangle:
            shape=CShape(name, shapetype, Left, Top, Width, Height, Fill, ws=self.ws)
        elif shapetype == msoShapeOval:
            shape=CShape(name, shapetype, Left, Top, Width, Height, Fill, ws=self.ws)
        elif shapetype == msoTextBox:
            shape=CShape(name, shapetype, Left, Top, Width, Height, Fill, Text=text,ws=self.ws)
        elif shapetype == msoOLEControlObject:
            shape=CShape(name, shapetype, Left, Top, Width, Height, Fill,ws=self.ws, Row=Row, Col=Col, control_dict=control_dict,Init_Value=Init_Value)
        self.ws.shapelist.append(shape)
        return shape # self.convTShape2CShape(shape)

    def AddLabel(self, TextOrientation=msoTextOrientationHorizontal, Left=0, Top=0 , Width=0, Height=0):
        shape=self.ws.addLabel( TextOrientation, Left, Top, Width, Height)
        return shape #self.convTShape2CShape(shape)
    
    def AddTextBox(self, Name="TextBox1", TextOrientation=msoTextOrientationHorizontal, Left=0, Top=0 , Width=0, Height=0,Text=""):
        #shape=self.ws.table.addLabel( TextOrientation, Left, Top, Width, Height)
        shapetype = msoOLEControlObject
        control_dict = {  "Name" : Name,
                            "BackColor"     : "#FFFFFF",
                            "BorderColor"   : "#000012",
                            "Caption"       : "",
                            "Height"        : Height,
                            "Left"          : Left,
                            "Col"           : None,
                            "Top"           : Top,
                            "Row"           : None,
                            "Type"          : "FormWindow",
                            "Visible"       : True,
                            "Width"         : Width,
                            "Components"    : [{"Name":Name,"Accelerator":"","BackColor":"#00000F","BorderColor":"#000006","BorderStyle":"fmBorderStyleNone","IconName":"",
                                               "Caption": Text,
                                               "Command": None ,"ControlTipText":"","ForeColor":"#000012","Height":Height,"Left":1,"Top":1,"Type":"TextBox","Visible":True,"Width":Width},
                                               ]}
        shape = self.ws.Shapes.AddShape(shapetype, Left, Top, Width, Height, Row=None, Col=None, Fill=0,text="",name=Name,control_dict=control_dict)
        #self.Buttonlist.append(self.shape)        
        
        return shape    
    
    def AddConnector(self, ConType, BeginX, BeginY, EndX, EndY):
        Left=BeginX
        Top=BeginY
        Width=EndX-BeginX
        Height=EndY-BeginY
        shape=self.ws.addConnector( ConType, Left, Top, Width, Height)
        return shape # self.convTShape2CShape(shape)
        
    
    def AddPicture(self, PicName, linktofile=False, savewithdocument=True, Left=0 , Top=0 , Width=0, Height=0):
        shape=self.ws.addPicture( PicName, Left, Top, Width, Height)
        return shape # self.convTShape2CShape(shape)
    
    def BuildFreeform(self, EditingType, X1, Y1):
        #print("BuildFreeForm:", X1, Y1)
        return CFreeForm(EditingType,X1,Y1)
        
    def Delete(self,shape):
        self.ws.deleteshape(shape)
        
    def Count(self):
        return len(self.ws.shapelist)
    
    def Range(self,input_array):
        searchname = str(input_array)
        searchname = searchname[2:-2] # remove []
        for Shape in self:
            if Shape.Name == searchname:
                return Shape
        return None
    
    def __iter__(self):
    #returning __iter__ object
        self.currentidx = 1
        self.currentshape = self.getShape(self.currentidx)
        self.stopiteration=False
        return self
    
    def __next__(self):
        self.currentshape = self.getShape(self.currentidx)
        if self.stopiteration or self.currentshape==None:
            raise StopIteration 
        self.currentidx += 1
        if self.currentidx > len(self.ws.shapelist):
            self.stopiteration=True       
        return self.currentshape

class CRangeList(list):
    def __init__(self, *args):
        list.__init__(self,*args)
        
    def get_count(self):
        return len(self)
    def set_count(self,range):
        pass        
    Count = property(get_count, set_count, doc='number of rows included in range')     
        
class CRange(str):
    def __new__(cls,t1=None,t2=None,ws=None,value="",parenttype="",parent=None):
        if ws==None:
            ws=ActiveSheet
        if t1==None and t2==None: #only type definition, only dummy needed
                return
        finished=False
        if type(t1) == vbclasses.Long:
            t1 = int(t1)
        if type(t2) == vbclasses.Long:
            t1 = int(t2)        
        if type(t1)==int and type(t2)==int:
            start=(t1,t2)
            end=(t1,t2)
            finished=True
        elif type(t1) == CRange and t2==None:
            #t1=t1.start
            start=t1.start
            end=t1.end
            finished=True
        elif type(t1)== CRange and type(t2) ==CRange:
            start=t1.start
            end=t2.end
            finished=True
        if not finished:
            if type(t1)== str:
                if IsNumeric(t1):
                    t1=int(t1)
                elif ":" in t1: # range as string a:b
                    cellsplit = t1.split(":")
                    temprng=CRange(cellsplit[0],cellsplit[1])
                    start= temprng.start
                    end  = temprng.end
                    finished=True
                else:
                    named_cell1 = ws.find_RangeName(t1,ws=ws)
                    if named_cell1 == None:
                        logging.debug("Named Range not found:", t1)
                        start=(1,1)
                        end=(1,1)
                        raise ValueError("CRange Error: Named Range not found:"+ t1)
                    else:
                        start=named_cell1.start
                        end  =named_cell1.end
                        if t2==None:
                            finished=True
        if not finished:
            if type(t1) == int and type(t2) == int:
                start=(t1,t2)
                end=(t1,t2)
                finished=True
            elif type(t2)==str:
                if IsNumeric(t2):
                    t2=int(t2)
                    start = (t1,t2)
                    end   = (t1,t2)
                    finished = True
                elif ":" in t2: # range as string a:b
                    cellsplit = t2.split(":")
                    temprng=CRange(cellsplit[0],cellsplit[1])
                    end  = temprng.end
                    finished=True
                else:
                    named_cell2 = ws.find_RangeName(t2,ws=ws)
                    if named_cell2 == None:
                        logging.debug("Named Range not found:", t1)
                        start=(1,1)
                        end=(1,1)
                        finished = True
                    else:
                        end  =named_cell2.end
                        finished = True
        if not finished:
            if type(t1) ==tuple and t2 == None: # t1 is a tuple (row,col)
                start=t1
                end  =t1
                finished=True
            elif type(t1)==tuple and type(t2)==tuple:
                start = t1
                end   = t2
                finished = True
        if not finished:
            if type(t1) == int and t2 == None:
                start = (t1, 1)
                end = (t1, 1)
                finished = True
        if not finished:
            raise ValueError("CRange: invalid type combination: t1="+str(type(t1))+" t2:"+str(type(t2)))
        row=start[0]
        col=start[1]
        max_cols = ws.tksheet.get_total_columns() #ColumnCount()
        max_rows = ws.tksheet.get_total_rows() #RowCount()
        if col <= max_cols and row <= max_rows:
            value = ws.tksheet.get_data(row-1, col-1) #getValueAt(row-1,col-1)
            if value == None:
                value = ""
        else:
            value = ""
        obj=str.__new__(cls,value)
        obj.ws = ws
        obj.Worksheet = obj.ws
        if parent==None:
            obj.Parent = obj.ws
        else:
            obj.Parent=parent
        obj.Interior = CInterior(parent=obj)
        obj.DisplayFormat = CDisplayFormat(parent=obj)
        obj.HasFormula = False
        obj.WrapText=False
        obj.Orientation=xlHorizontal
        obj.start = start
        obj.end   = end
        obj.Formula=""
        obj.ParentType=parenttype
        obj.iter_start=None
        obj.currentcell=obj.start
        obj.CountLarge = (obj.end[0]-obj.start[0]+1)*(obj.end[1]-obj.start[1]+1)
        obj.Font = CFont(obj.Parent.default_font[0],obj.Parent.default_font[1], parent=obj)
        obj.default_font = obj.Parent.default_font
        if obj == None:
            logging.debug("CRange -New: Error in Obnject creation")
        return obj

    # Attributes
    def dummy(self,value):
        pass
    
    def get_address(self):
        return self.start
    def set_address(self,range):
        pass
    Address = property(get_address, set_address, doc='adress of first cell in range')

    def set_cells(self,rowlist):
        pass
    def get_cells(self):
        if self.Parent.ParentType=="Columns":
            startrow=self.Parent.Parent.start[0]
            lastrow =self.Parent.Parent.end[0]
            column = self.start[1]
            return CRange((startrow,column),(lastrow,column),parent=self,ws=self.ws)
        else:
            logging.debug("get_cells-Error")
    Cells = property(get_cells, set_cells, doc='rows included in range')   

    def get_column(self):
        #row = CRow(self.start[0])
        return self.start[1]
    def set_column(self,range):
        pass 
    Column = property(get_column, set_column, doc='row of first cell in range')

    def set_columns(self,rowlist):
        pass
    def get_columns(self):
        return CRange((self.start[0],self.start[1]),(self.start[0],self.end[1]),ws=self.ws,parent=self,parenttype="Columns")
    Columns = property(get_columns, set_columns, doc='Columns in Range')

    def get_columnwidth(self):
        #colname=self.ws.tablemodel.getColumnName(self.start[1]-1)
        #width=self.ws.tablemodel.columnwidths[colname]
        width = self.ws.tksheet.column_width(self.start[1]-1)
        return width
    def set_columnwidth(self, width):
        for col in range(self.start[1],self.end[1]+1):
            #colname=self.ws.tablemodel.getColumnName(col-1)
            #self.ws.tablemodel.columnwidths[colname] = width
            self.ws.tksheet.column_width(col-1, int(width))
        return
    ColumnWidth = property(get_columnwidth, set_columnwidth, doc='ColumnWidth of CRange')

    def get_count(self):
        return (self.end[0]-self.start[0]+1)*(self.end[1]-self.start[1]+1)
    def set_count(self,p_range):
        pass         
    Count = property(get_count, set_count, doc='number of cells in range')
      
    def get_entirecolumn(self):
        entirecolumn = CRange((1,self.start[1]),(self._get_LastRow(),self.end[1]),parenttype="EntireColumn")
        return entirecolumn
    def set_entirecolumn(self,param):
        pass
    EntireColumn = property(get_entirecolumn, set_entirecolumn, doc='entirecolumn of range')

    def get_entirerow(self):
        entirerow = CRange((self.start[0],1),(self.end[0],self._get_LastColumn()),parenttype="EntireRow")
        return entirerow
    def set_entirerow(self,param):
        pass
    EntireRow = property(get_entirerow, set_entirerow, doc='entirerow of range')

    def _get_Height(self):
        x1,y1,x2,y2 = self.ws.getCellCoords(self.Row-1,self.Column-1)
        return (y2-y1)
    Height  = property(_get_Height, dummy, doc='Height of cell')
    
    def get_Hidden(self):
        if self.ParentType=="EntireRow":
            #return self.start[0]-1 in ActiveSheet.table.hiderowslist
            #return not (self.start[0]-1 in ActiveSheet.tksheet.displayed_rows)
            return ActiveSheet.is_row_hidden(self.start[0]-1)
        elif self.ParentType=="EntireColumn":
            return False
        else:
            return False
    def _set_hiderow(self,row,newval):
        ActiveSheet.set_hiderow(row, newval)
    def _set_hidecolumn(self,col,newval):
        ActiveSheet.set_hidecolumn(col, newval)
    def set_Hidden(self,hide):
        if self.ParentType=="EntireRow":
            for row in range(self.start[0],self.end[0]+1):
                self._set_hiderow(row,hide)            
        elif self.ParentType=="EntireColumn":
            for col in range(self.start[1],self.end[1]+1):
                    self._set_hidecolumn(col,hide)
        else:
            return False
    Hidden = property(get_Hidden, set_Hidden, doc='hidden attribute of range')

    def _get_Left(self):
        x1,y1,x2,y2 = self.ws.getCellCoords(self.Row-1,self.Column-1)
        return x1    
    Left = property(_get_Left, dummy, doc='Left coordinates of cell')

    def get_row(self):
        return self.start[0]
    def set_row(self,range):
        pass
    Row = property(get_row, set_row, doc='row of first cell in range')
    
    def _set_rowheight(self,row,newval):
        #ActiveSheet.tksheet.rowheightlist[row-1]=newval
        newval = int(newval)
        displayed_row = ActiveSheet.data_row_to_displayed(row-1)
        ActiveSheet.tksheet.row_height(row=displayed_row, height=newval)
        ActiveSheet.Redraw_table()
    def _get_rowheight(self,row):
        return ActiveSheet.tksheet.row_height(row=row-1, height=None)
        #return ActiveSheet.tksheet.rowheightlist[row-1]    
    def set_RowHeight(self, newval):
        newval = newval * xlvp2py_guifactor
        for row in range(self.start[0], self.end[0]+1):
            self._set_rowheight(row,newval)
    def get_RowHeight(self):
        #rowheight = ActiveSheet.tksheet.rowheightlist.get(self.start[0]-1,ActiveSheet.tksheet.default_rowheight)
        rowheight = ActiveSheet.tksheet.row_height(row=self.start[0]-1, redraw=False)
        rowheight = int(rowheight/xlvp2py_guifactor)
        return rowheight
    RowHeight = property(get_RowHeight, set_RowHeight, doc='CRange RowHeight')

    def get_rows(self):
        return CRange((self.start[0],self.start[1]),(self.end[0],self.start[1]),ws=self.ws,parent=self,parenttype="Rows")
    def set_rows(self,rowlist):
        pass
    Rows = property(get_rows, set_rows, doc='rows included in range')
    
    def get_text(self):
        return self.get_value()
    def set_text(self,value):
        self.set_value(value)
    Text = property(get_text, set_text, doc='value of Textvalue of Range')
    
    def _get_Top(self):
        x1,y1,x2,y2 = self.ws.getCellCoords(self.Row-1,self.Column-1)
        return y1
    Top  = property(_get_Top, dummy, doc='Top coordinates of cell')
  
    def get_value(self):
        row=self.start[0]
        col=self.start[1]
        endrow = self.end[0]
        endcol = self.end[1]
        if self.ws.tksheet == None:
            self.ws.tksheet = ActiveSheet.tksheet
        max_cols = self.ws.tksheet.get_total_columns() #ColumnCount()
        max_rows = self.ws.tksheet.get_total_rows() #RowCount()()
        if row != endrow or col != endcol:
            if col <= max_cols and endcol <= max_cols and row <= max_rows and endrow <= max_rows:
                value = self.ws.tksheet.get_data(row-1,col-1, endrow, endcol)
            else:
                value = ""
        else:
            if col <= max_cols and row <= max_rows:
                value = self.ws.tksheet.get_data(row-1,col-1)
            else:
                value = ""
        return value
    def set_value(self,value):
        row=self.start[0]
        col=self.start[1]        
        if self.ws.tksheet == None:
            self.ws.tksheet = ActiveSheet.tksheet
        if row >= self.ws.tksheet.get_total_rows(): #RowCount()():
            #add rows
            startrow = self.ws.tksheet.get_total_rows() #RowCount()()-1
            keys_Added = self.ws.tksheet.insert_rows(row-startrow+1)
        if col >self.ws.tksheet.get_total_columns(): #getColumnCount():
            #add columnsrows
            pass
        if row <=self.ws.tksheet.get_total_rows() and col <=self.ws.tksheet.get_total_columns(): #getColumnCount():
            curval = self.ws.tksheet.get_data(row-1, col-1)
            if curval != value:
                self.ws.tksheet.span(row-1, col-1).data = value   # (value, row-1, col-1)
                self.ws.tksheet.event_row=row-1
                self.ws.tksheet.event_col=col-1
                self.ws.Synch_Evt_SheetChange(self.ws.tksheet)
    Value = property(get_value, set_value, doc='value of first Cell in Range')
    
    def _get_Width(self):
        x1,y1,x2,y2 = self.ws.getCellCoords(self.Row-1,self.Column-1)
        return x2-x1
    Width = property(_get_Width, dummy, doc='Width of cell')

# Methods
    def Activate(self):
        for cell in self:
            row=cell.Row
            col=cell.Column
            self.ws.tksheet.select_cell(row-1, col-1, redraw=True)

    def CellsFct(self,row,column):
        return CRange(row,column, ws=self.ws)    

    def ClearContents(self):
        self.ws.tksheet.span(self.start,self.end).clear()
        # for cell in self:
        #    cell.Value=""
            
    def Copy(self,dst):
        #calculate offset for copy
        offset_row = dst.Row-self.start[0]
        offset_col = dst.Column-self.start[1]
        if offset_col>0:
            for row in range(self.end[0],self.start[0]-1,-1):
                for col in range(self.end[1],self.start[1]-1,-1):
                    cell=CRange(row,col,ws=self.ws)
                    cell.offset(offset_row,offset_col).Value=cell.Value
        else:
            for row in range(self.start[0],self.end[0]+1):
                for col in range(self.start[1],self.end[1]+1):
                    cell=CRange(row,col,ws=self.ws)
                    cell.offset(offset_row,offset_col).Value=cell.Value            
            
    def End(self,param):
        row=self.start[0]
        column=self.start[1]
        if param==xlToRight:
            column=-1
            for cell in self:
                if cell.Value!="":
                    if cell.Column>column:
                        column=cell.Column
            if column==-1:
                column = self.end[1]+2 #I do not knbow if this is correct, but the column must be higher than end +1 to show that the row is empty
        elif param==xlUp:
            logging.debug("End -xlToUp")
            
        elif param==xlDown:
            logging.debug("End -xlDown")
        else: # xlToLeft
            logging.debug("End -xlToLeft")
        return CRange(row,column)    
                
    def Find(self,What="", after=None, LookIn=xlFormulas, LookAt= xlWhole, SearchOrder=xlByRows, SearchDirection=xlNext, MatchCase= True, SearchFormat= False):
        if after!=None:
            self.iter_start = after.start
        else:
            self.iter_start = None
        #if self.ParentType=="Rows":
        #    if self.end[1]==1:
        #        self.end=(self.end[0],self._get_LastColumn())  # search in complete row
        for cell in self:
            cell_val=str(cell.Value)
            if "§" in cell_val:
                special_char = "§" # remove comments in excel sheet
                index = cell_val.find(special_char)
                if index != -1:
                    cell_val = cell_val[:index]            
            if cell_val == What:
                return cell
        return None
    
    def Insert(self, Shift=0, CopyOrigin=0):
        if self.ws.tksheet == None:
            self.ws.tksheet = ActiveSheet.tksheet
        if Shift== xlDown and CopyOrigin== xlFormatFromLeftOrAbove:
            self.ws.tksheet.insert_rows(rows=self.end[0]-self.start[0]+1, idx=self.start[0]-1)
        else:
            logging.debug("Insert: unknown parameters")
            
    def Offset(self, offset_row, offset_col):
        return self.offset(offset_row, offset_col)    
            
    def offset(self, offset_row, offset_col):
        newstart=(self.start[0]++offset_row,self.start[1]+offset_col)
        newend=(self.end[0]++offset_row,self.end[1]+offset_col)
        new_range=CRange(newstart,newend,ws=self.ws)
        return new_range
    
    def Select(self):
        self.ws.setSelectedCells(self.start[0]-1, self.end[0]-1,self.start[1]-1,self.end[1]-1)
        
    def check_if_empty_row(self):
        if self.ws.tksheet == None:
            self.ws.tksheet = ActiveSheet.tksheet
        if self.Row > self.ws.tksheet.get_total_rows(): #RowCount()():
            return True
        r=self.Row-1
        cols = self.ws.tksheet.get_total_columns() #getColumnCount()
        for c in range(0,cols):
            #absr = self.get_AbsoluteRow(r)
            val = self.ws.tksheet.get_data(r,c)
            if val != None and val != '':
                return False
        return True

    def __iter__(self):
    #returning __iter__ object
        if self.iter_start!=None:
            self.currentcell = self.iter_start
        else:
            self.currentcell = self.start
        self.stopiteration=False
        return self
    
    def __next__(self):
        currentresult = CRange(self.ws.Cells(self.currentcell[0],self.currentcell[1]),parent=self,ws=self.ws)
        if currentresult == None:
            raise StopIteration
        if self.stopiteration:
            raise StopIteration
        currentcell_col=self.currentcell[1]
        currentcell_row=self.currentcell[0]
        currentcell_col+=1
        if currentcell_col>self.end[1]:
            currentcell_row+=1
            currentcell_col = self.start[1]
            if currentcell_row > self.end[0]:
                self.stopiteration=True
        self.currentcell = (currentcell_row,currentcell_col)
        return currentresult
    
    def __eq__(self,other):
        if type(other) == int:
            value=self.get_value()
            if type(value)==int:
                intvalue=value
                fn_return_value=(intvalue==other)
            else:
                if value.isnumeric():
                    intvalue=int(value)
                    fn_return_value=(intvalue==other)
                else:
                    fn_return_value=False
            return fn_return_value
        else:
            value=self.get_value()
            strvalue=str(value)
            strother=str(other)    
            fn_return_value=(strvalue==strother)
            return fn_return_value            

    def __lt__(self,other):
        if type(other) == int:
            value=self.get_value()
            if type(value)==int:
                intvalue=value
                fn_return_value=(intvalue<other)
            else:
                if value.isnumeric():
                    intvalue=int(value)
                    fn_return_value=(intvalue<other)
                else:
                    fn_return_value=False
            return fn_return_value
        value=str(self)
        fn_return_value=(value<other)
        return fn_return_value
        
    def __gt__(self,other):
        if type(other) == int:
            value=self.get_value()
            if type(value)==int:
                intvalue=value
                fn_return_value=(intvalue>other)
            else:
                if value.isnumeric():
                    intvalue=int(value)
                    fn_return_value=(intvalue>other)
                else:
                    fn_return_value=False
            return fn_return_value
        value=str(self)
        fn_return_value=(value>other)
        return fn_return_value
    
    def __le__(self,other):
        if type(other) == int:
            value=self.get_value()
            if type(value)==int:
                intvalue=value
                fn_return_value=(intvalue<=other)
            else:
                if value.isnumeric():
                    intvalue=int(value)
                    fn_return_value=(intvalue<=other)
                else:
                    fn_return_value=False
            return fn_return_value

        value=str(self)
        fn_return_value=(value<=other)
        return fn_return_value
        
    def __ge__(self,other):
        if type(other) == int:
            value=self.get_value()
            if type(value)==int:
                intvalue=value
                fn_return_value=(intvalue>=other)
            else:
                if value.isnumeric():
                    intvalue=int(value)
                    fn_return_value=(intvalue>=other)
                else:
                    fn_return_value=False
            return fn_return_value

        value=str(self)
        fn_return_value=(value>=other)
        return fn_return_value    
        
        
    def __int__(self):
        value=self.get_value()
        return int(value)
    
    def __str__(self):
        value=self.get_value()
        return str(value)
    
    def __pow__(self, other):
        value = str(self)
        if value.isnumeric():
            intvalue=int(value)
            return intvalue ** other
        else:
            return 1
    
    def __rpow__(self, other):
        value = str(self)
        if value.isnumeric():
            intvalue=int(value)
            return other ** intvalue
        else:
            return other
        
    def __add__(self,other):
        value = str(self)
        if not isinstance(other,str):
            if value.isnumeric():
                intvalue=int(value)
                return other + intvalue
            else:
                return other
        else:
            return str(self) + str(other)
        
    def __radd__(self,other):
        value = str(self)
        if not isinstance(other,str):
            if value.isnumeric():
                intvalue=int(value)
                return other + intvalue
            else:
                return other
        else:
            return str(other) + value
        
    def __sub__(self,other):
        value = str(self)
        if value.isnumeric():
            intvalue=int(value)
            return intvalue - other
        else:
            return -other
        
    def __rsub__(self,other):
        value = str(self)
        if value.isnumeric():
            intvalue=int(value)
            return  other -intvalue
        else:
            return other
        
    def __mul__(self,other):
        value = str(self)
        if value.isnumeric():
            intvalue=int(value)
            return intvalue * other
        else:
            return other
        
    def __rmul__(self,other):
        value = str(self)
        if value.isnumeric():
            intvalue=int(value)
            return intvalue * other
        else:
            return other

# other functions    
    def upper(self):
        value=str(self)
        return value.upper()
    
# internal functions
    def _get_LastRow(self):
        return self.ws.tksheet.get_total_rows() #RowCount()()
    
    def _get_LastColumn(self):
        return self.ws.tksheet.get_total_columns() #getColumnCount()
    
    def last_non_zero_index(self, l):
        result = 0
        i = len(l) - 1
        while i >= 0:
            if l[i] != "" and type(l[i]) != list:
                result = i
                break
            i -= 1
        return result
    
    def get_last_used_column(self, lastrow=None):
        last_used_column_max = 0
        if lastrow == None:
            lastrow = self.end[0]
        for row in range(self.start[0], lastrow+1):
            row_content = self.ws.tksheet[row-1:row].data
            last_used_column = self.last_non_zero_index(row_content)
            if last_used_column > last_used_column_max:
                last_used_column_max = last_used_column
        return last_used_column_max
            
    
class CShapeRange(object):
    def __init__(self,ws=None):
        if ws==None:
            ws=ActiveSheet
        self.ws=ws
        self.shapes=ws.getSelectedShapes()
        self.Line=CLine(shape=self.shapes)
        self.Fill=CFill(shape=self.shapes)
        self.ws=ws
        
    def __getitem__(self, k):
        #print("Getitem",k)
        self.shapes=self.ws.getSelectedShapes()
        if k-1<len(self.shapes):
            return self.shapes[k-1]
        else:
            return None
         
    def __setitem__(self,k,value):
        self.ws.setSelectedShapes(value)
        
    def set_Name(self, newval):
        for shape in self.shapes:
            shape.Name=newval
    def get_Name(self):
        return self.shapes[0].Name
    Name = property(get_Name, set_Name, doc='ShapeRange-Name')
    
class CColumns(object):
    def __init__(self,startcol=None,lastcol=None,parent=None,ws=None):
        self.ws=ws
        if startcol==None:
            self.startcol=parent.start[1]
            self.lastcol =parent.end[1]
        else:
            self.startcol = startcol
            self.lastcol  = lastcol
        self.parent=parent
        self.currentcol = startcol
        self.Column = startcol
        #self.Cells = CRange(self.parent.start[0],self.currentcol),(self.parent.end[0],self.currentcol),parent=self,ws=self.ws)
        
    def get_Count(self):
        return self.lastcol-self.startcol
    def set_Count(self,value):
        pass
    Count = property(get_Count,set_Count,doc="Selection Count")
    
    #python substitute function
            
    def __iter__(self):
    #returning __iter__ object
        self.currentcol = self.startcol
        self.stopiteration=False
        return self
    
    def __next__(self):
        self.Column = self.currentcol
        currentresult = CRange((self.parent.start[0],self.currentcol),(self.parent.end[0],self.currentcol),parent=self,ws=self.ws)
        if self.stopiteration:
            raise StopIteration
        self.currentcol+=1
        if self.currentcol>self.lastcol: 
            self.stopiteration=True
        return currentresult    
   
class CSelection(object):
    def __init__(self,selrange=None,ws=None):
        if ws==None:
            ws=ActiveSheet
        self.ws=ws
        if self.ws == None:
            self.selectedrange = None
        else:
            if selrange != None:
                self.selectedrange=selrange
            else:
                self.selectedrange = CRange(self.ws.Selection)
        
    def set_ColRange(self, newval):
        pass
    def get_ColRange(self):
        selectedcol=ActiveSheet.getSelectedColumn() + 1
        if selectedcol != None:
            colrange=(selectedcol,selectedcol+1)
        else:
            colrange = (0, 0)
        return colrange
    colrange = property(get_ColRange, set_ColRange, doc='ColRange')
    
    def set_Column(self, newval):
        pass
    def get_Column(self):
        selectedcol=ActiveSheet.getSelectedColumn()+1
        return selectedcol
    Column = property(get_Column, set_Column, doc='Column')
    
    def set_Columns(self, newval):
        pass
    def get_Columns(self):
        selectedcol=ActiveSheet.getSelectedColumn()+1
        return CColumns(selectedcol,selectedcol+1,ws=self.ws,parent=self)
    Columns = property(get_Columns, set_Columns, doc='Columns')
    
    def set_Cells(self, newval):
        pass
    def get_Cells(self):
        selectedcol=ActiveSheet.getSelectedColumn()+1
        selectedrow=ActiveSheet.getSelectedRow()+1
        return CRange(selectedrow,selectedcol)
    Cells = property(get_Cells, set_Cells, doc='Selection-Cells')    
    
    def _selectedRow(self):
        return [self.ws.getSelectedRow()]
    
    def get_entirerow(self):
        selectedrows = self._selectedRow()
        lastcolumn =  self.ws.tksheet.get_total_columns()
        entirerow = CRange((selectedrows[0],1),(selectedrows[-1],lastcolumn),parenttype="EntireRow")
        return entirerow
    def set_entirerow(self,param):
        pass
    EntireRow = property(get_entirerow, set_entirerow, doc='entirerow of selection')
    
    def set_Row(self, newval):
        pass
    def get_Row(self):
        return ActiveSheet.getSelectedRow()+1
    Row = property(get_Row, set_Row, doc='Selection-Row')    
    
    def Rows(self):
        selectedrows = self._selectedRow()
        selectedcrows = []
        for row in selectedrows:
            selectedcrows.append(CRow(row))
        return selectedcrows

    def set_RowRange(self, newval):
        pass
    def get_RowRange(self):
        selectedrows=self.EntireRow
        if selectedrows != []:
            rowrange=(selectedrows.start[0],selectedrows.end[0]+1)
        return rowrange
    rowrange = property(get_RowRange, set_RowRange, doc='RowRange')
    
    def Select(self):
        self.ws.setSelectedCells(self.selectedrange.start[0]-1, self.selectedrange.end[0]-1,self.selectedrange.start[1]-1,self.selectedrange.end[1]-1)
        pass
    
    def set_ShapeRange(self, newval):
        pass
    def get_ShapeRange(self):
        return CShapeRange(ws=self.ws)
    ShapeRange = property(get_ShapeRange, set_ShapeRange, doc='ShapeRange')

class CRow(object):
    def __init__(self,rownumber,rowrange=None):
        self.Row = rownumber
        self.Rowrange=rowrange
        self.EntireRow = CEntireRow(rownumber,rowrange)
        self.Hidden_value = False
        self.Top = 0  #*HL
        self.Height = 12 #*HL
        
    def Select(self):
        ActiveSheet.tksheet.gotoCell(self.Row-1,0)
        return
    
    def _set_hiderow(self,row,newval):
        ActiveSheet.set_hiderow(row, newval)
        ActiveSheet.Redraw_table()
                
    def _set_rowheight(self,row,newval):
        #ActiveSheet.tksheet.rowheightlist[row-1]=newval
        newval = int(newval)
        displayed_row = ActiveSheet.data_row_to_displayed(row-1)
        ActiveSheet.tksheet.row_height(row=displayed_row, height=newval)
        ActiveSheet.Redraw_table()
        
    def _get_rowheight(self,row):
        return ActiveSheet.tksheet.row_height(row=row-1, height=None)
        #return ActiveSheet.tksheet.rowheightlist[row-1]
        
    def set_Hidden(self, newval):
        self.Hidden_value = newval
        if self.Rowrange:
            for row in self.Rowrange:
                self._set_hiderow(row,newval)
        else:
            self._set_hiderow(self.Row,newval)
    
    def get_Hidden(self):
        return ActiveSheet.is_row_hidden(self.Row-1)
        
    Hidden = property(get_Hidden, set_Hidden, doc='Hiddenvalue of CRow')
    
    def set_RowHeight(self, newval):
        newval = newval * xlvp2py_guifactor
        if self.Rowrange:
            for row in self.Rowrange:
                self._set_rowheight(row,newval)
        else:
            self._set_rowheight(self.Row,newval)
    
    def get_RowHeight(self):
        #rowheight = ActiveSheet.tksheet.rowheightlist.get(self.Row-1,ActiveSheet.tksheet.default_rowheight)
        rowheight = ActiveSheet.tksheet.row_height(row=self.Row-1, redraw=False)
        rowheight = int(rowheight/xlvp2py_guifactor)
        return rowheight
        
    RowHeight = property(get_RowHeight, set_RowHeight, doc='Hiddenvalue of CRow')
        
class CEntireRow(object):
    def __init__(self,rownumber,colrange=None,ws=None):
        self.Rownumber = rownumber
        self.Hidden_value = False
        self.Colrange=colrange
        self.Columns=CColumns(CRange((self.Rownumber,1),colrange,ws=ws))
        
    def _set_hiderow(self,row,newval):
        ActiveSheet.set_hiderow(row, newval)
                
    def set_Hidden(self, newval):
        self.Hidden_value = newval
        self._set_hiderow(self.Rownumber,newval)
    
    def get_Hidden(self):
        return ActiveSheet.is_row_hidden(self.Rownumber-1)
        
    Hidden = property(get_Hidden, set_Hidden, doc='Hiddenvalue of CRow')
    
    def set_value(self, newval):
        newval = newval * xlvp2py_guifactor
        if self.Rowrange:
            for row in self.Rowrange:
                self._set_rowheight(row,newval)
        else:
            self._set_rowheight(self.Row,newval)
    
    def get_value(self):
        row_value = ActiveSheet.tksheet[self.Rownumber]
        return row_value
        
    Value = property(get_value, set_value, doc='Entirerow Value')    
        
class CColumn(object):
    def __init__(self,colnumber,rowrange=None):
        self.Column=colnumber
        self.EntireColumn = CEntireColumn(colnumber)
        self.columnwidth = 5
        self.Hidden = False
        self.rowrange=rowrange
        #self.Cells = CRange(rowrange,(colnumber,colnumber))
        
    def get_columnwidth(self):
        return self.columnwidth
    
    def set_columnwidth(self, value):
        value = int(value)
        self.columnwidth = value
        ActiveSheet.tksheet.column_width(self.Column-1,value)
        
    def get_cells(self):
        cell_list=[]
        if self.rowrange:
            for cellrow in self.rowrange:
                cell_list.append(Cells(cellrow,self.Column))
        
        return cell_list
    
    def set_cells(self, value):
        pass
    
    def _set_hidecolumn(self,col,newval):
        ActiveSheet.set_hidecolumn(col, newval)    
        
    ColumnWidth = property(get_columnwidth, set_columnwidth, doc='Column Width')
    
    Cells = property(get_cells, set_cells, doc='Cells in Column')
        
#*******************************************
# Internal Object EntireColumn - should not be used anymore
#*******************************************            
class CEntireColumn(object):
    def __init__(self,colnumber):
        self.Columnnumber = colnumber
        self.Hidden = False

#*******************************************
# Excel Object Interior https://learn.microsoft.com/en-us/office/vba/api/excel.interior(object)
#*******************************************            
class CInterior(object):
    def __init__(self,parent=None):
        self.parent=parent
        
    def get_color(self):
        if self.parent == None:
            return tkcolor2int("#FFFFFF")
        try:
            if type(self.parent) == CRange:
                colorvalue = self.parent.ws.tksheet.getColorAt(self.parent.start[0]-1,self.parent.start[1]-1)
                if colorvalue==None:
                    colorvalue="#FFFFFF"
                return tkcolor2int(colorvalue)
            else:
                Debug.Print("CInterior Type ERROR")
                return tkcolor2int("#FFFFFF")
        except BaseException as e:
            logging.debug(e, exc_info=True) 
            return tkcolor2int("#FFFFFF")
    
    def set_color(self, value):
        if self.parent==None:
            return
        if type(value) == int:
            valuestr=int2tkcolor(value)
        elif type(value) == str:
            if value.isnumeric():
                valuestr=int2tkcolor(int(value))
            else:
                valuestr=value
        else:
            valuestr=f'#{value[0]:02x}{value[1]:02x}{value[2]:02x}'
        if type(self.parent) == CRange:
            for cell in self.parent:
                self.parent.ws.tksheet.highlight((cell.start[0]-1,cell.start[1]-1), bg=valuestr)
    Color = property(get_color, set_color, doc='Cell Color')    

#*******************************************
# Excel Object DisplayFormat https://learn.microsoft.com/en-us/office/vba/api/excel.displayformat
#*******************************************            
class CDisplayFormat(object):
    def __init__(self,parent=None):
        self.parent=parent
        self.Interior = CInterior(parent=parent)

#*******************************************
# Excel Object Color https://learn.microsoft.com/en-us/office/vba/api/excel.colorformat
#*******************************************            
class CColor(object):
    def __init__(self,color,shape=None,tksheet=None,line=None,cell=None,fg=False):
        self.line=line
        self.shape=shape
        self.tksheet=tksheet
        self.cell=cell
        self.color = color # int 0xrrggbb
        self.r = 0
        self.g = 0
        self.b = 0
        
    def tkcolor(self):
        return f'#{self.r:02x}{self.g:02x}{self.b:02x}'
    
    def get_rgb(self):
        colorvalue=self.color
        if self.cell!=None:
            colorvalue_str = self.tksheet.getColorAt(self.cell.Row-1,self.cell.Column-1)
            colorvalue=tkcolor2rgb(colorvalue_str)
            #print("get_rgb:"+self.cell.Name+"-"+colorvalue_str)
        elif self.shape!=None:
            colorvalue=self.color
        return colorvalue
    
    def set_rgb(self, value):
        if type(value) == int:
            valuestr=int2tkcolor(value)
        elif type(value) == str:
            valuestr=value
        else:
            valuestr=f'#{int(value[0]):02x}{int(value[1]):02x}{int(value[2]):02x}'
        if self.cell!=None:
            self.cell.tksheet.setColorAt(self.cell.Row-1,self.cell.Column-1, valuestr)
            self.rgb_val=valuestr
        if self.shape!=None:
            if type(self.shape)==list:
                if self.shape==[]:
                    shape=None
                else:
                    shape=self.shape[0]
            else:
                shape=self.shape
            if shape!=None:
                shape.Fillcolor=valuestr
                if shape.rectidx!=0:
                    shape.Tablecanvas.itemconfig(shape.rectidx,fill=valuestr)
                else:
                    shape.updatecontrol = True
                    #shape.Worksheet.drawShape(shape)
                    shape.updatecontrol = False
                
                #shape.Tshape.Fillcolor=valuestr
                #print("set_rgb:"+self.shape.Name+"-"+valuestr)
            else:
                logging.debug("set_rgb: Shape not found" +str(self.shape)+ "-"+valuestr)                         
        else:
            logging.debug("set_rgb: Shape not found" + "-"+valuestr)                
                
        self.rgb_val=valuestr
        rgb=tkcolor2rgb(valuestr)
        self.r = rgb[0]
        self.g = rgb[1]
        self.b = rgb[2]
        
    rgb = property(get_rgb, set_rgb, doc='Color RGB')
    
    RGB = property(get_rgb, set_rgb, doc='Color RGB') 

#*******************************************
# Excel Object Line https://learn.microsoft.com/en-us/office/vba/api/excel.lineformat
#*******************************************        
class CLine(object):
    def __init__(self,linecolor=0,shape=None):
        self.Weight = 8
        self.ForeColor = CColor((255,0,0),line=True,shape=shape,fg=True)
        self.BackColor = CColor((255,0,0),line=True,shape=shape,fg=False)
        self.Visible = True
        self.Transparency = 0
        self.shape=shape
        self.BeginArrowheadStyle_val = 0
        self.EndArrowheadStyle_val= 0
        
    def get_EndArrowheadStyle(self):
        return self.EndArrowheadStyle_val
    
    def set_EndArrowheadStyle(self,newval):
        self.EndArrowheadStyle_val=newval
        if type(self.shape)==list:
            for shape in self.shape:
                shape.EndArrowheadStyle_val=newval
        else:
            self.shape.EndArrowheadStyle_val=newval      
        
    EndArrowheadStyle = property(get_EndArrowheadStyle,set_EndArrowheadStyle,doc="EndArrowheadStyle")
    
    def get_BeginArrowheadStyle(self):
        return self.BeginArrowheadStyle_val
    
    def set_BeginArrowheadStyle(self,newval):
        self.BeginArrowheadStyle_val=newval
        if type(self.shape)==list:
            for shape in self.shape:
                shape.BeginArrowheadStyle_val=newval
        else:
            self.shape.BeginArrowheadStyle_val=newval      
        
    BeginArrowheadStyle = property(get_BeginArrowheadStyle,set_BeginArrowheadStyle,doc="BeginArrowheadStyle")    
    
#*******************************************
# Excel Object Fill https://learn.microsoft.com/en-us/office/vba/api/excel.fillformat
#*******************************************        
class CFill(object):
    def __init__(self,Fill="",shape=None):
        self.Visible = True
        self.ForeColor = CColor((0, 0, 0),shape=shape,fg=True)
        self.Transparency_val = 0
        self.shape=shape
        
#***************************
# Property Transparency https://learn.microsoft.com/en-us/office/vba/api/excel.fillformat.transparency
    def get_transparency(self):
        return self.Transparency_val
    
    def set_transparency(self,newval):
        self.Transparency_val=newval
        if newval >=0.5:
            zorder=1 #to Bottum
        else:
            zorder=0 # toTop
        if type(self.shape)==list:
            for shape in self.shape:
                shape.ZOrder_Val=zorder
        else:
            self.shape.ZOrder_Val=zorder      
        
    Transparency = property(get_transparency,set_transparency,doc="Transparency")
        
    #***************************
    # Method Solid https://learn.microsoft.com/en-us/office/vba/api/excel.fillformat.transparency
    def Solid(self):
        pass

#*******************************************
# Excel Object Shape https://learn.microsoft.com/en-us/office/vba/api/excel.shape
#*******************************************        
class CShape(object):
    def __init__(self, name, shapetype, Left, Top, Width=0, Height=0, Fillcolor=0, Text="",Row=None, Col=None, rotation=0,orientation=msoTextOrientationHorizontal,AlternativeText="",ws=None,nodelist=None,control_dict=None,Init_Value=None,linecolor=None,zorder=0,segmenttype=msoSegmentLine,editingtype=msoEditingAuto):
        filltkcolor=int2tkcolor(Fillcolor)
        if linecolor:
            linetkcolor=int2tkcolor(linecolor)
        else:
            linetkcolor=None
        if shapetype==msoOLEControlObject:
            if type(Width)==int:
                Width = Width *1.55
                Height = Height *1.55
        if ws==None:
            ws=ActiveSheet
        #if tshape==None:
        #    tshape=CTShape(name, shapetype, Left, Top, Width, Height, FillTKcolor=filltkcolor, LineTKcolor=linetkcolor,Text=Text,Row=Row, Col=Col,rotation=rotation,orientation=orientation,nodelist=nodelist,ws=ws, tablecanvas=ws.tksheet,control_dict=control_dict,Init_Value=Init_Value,zorder=zorder,segmenttype=segmenttype,editingtype=editingtype)
        #self.Tshape=tshape 
        self.Shapetype = shapetype
        if shapetype in (msoShapeRectangle, msoShapeOval):
            self.Type = msoAutoShape
            self.AutoShapeType = shapetype
        else:
            self.Type = shapetype
        self.Tablecanvas=ws.tksheet.MT
        if ws==None:
            ws=ActiveSheet
        self.Worksheet=ws
        #self.Tshape=tshape
        self.Name_val=name
        self.rectidx=0
        self.textidx=0
        self.tablename = name #test ws.table.tablename
        self.Left_val = Left
        self.Active = True
        self.Top_val = Top
        self.formwin=None
        self.Width = Width
        self.Height = Height
        self.Row=Row
        self.Col=Col
        self.Rotation = rotation
        self.Visible_val = True
        if Top==None:

            self.TopLeftCell_Row=Row # -1 removed 13.5.2024 HL
            self.TopLeftCell_Col=Col # -1 removed 13.5.2024 HL
        else:
            self.TopLeftCell_Row=ws.calc_row_from_y(Top+1) + 1
            self.TopLeftCell_Col=ws.calc_col_from_x(Left+1) + 1
        self.Index = 0
        self.OnAction_val = ""
        self.ZOrder_Val=zorder
        self.Locked = False
        self.Text = Text
        self.TextFrame2 = CTextFrame2(text=Text,shape=self)
        self.AlternativeText = AlternativeText
        self.Fill = CFill(Fillcolor,shape=self)
        self.Line = CLine(linecolor,shape=self)
        self.control_dict = copy.deepcopy(control_dict)
        self.OnAction = ""

        self.updatecontrol=False
        #self.tablename = tablecanvas.name
        self.image=None
        self.AlternativeText=""
        self.control_dict = copy.deepcopy(control_dict)
        self.Init_Value=Init_Value
        self.Fillcolor = filltkcolor
        self.LineColor = linetkcolor
        self.nodelist=nodelist
        self.SegmentType = segmenttype
        self.EditingType = editingtype
        self.format_dict = {}
        #self.Worksheet.shapelist.append(self)        

        #if tshape==None:    # create new tshape
        #    self.Tshape=self.Tablecanvas.addshape(name, shapetype, Left, Top, Width, Height, Fillcolor, Text="",rotation=0,orientation=1,tablecanvas=None)
            
    def updateShape(self,Fill=None,Text=None, AltText=None):
        if Fill:
            fillcolor=Fill.Color
        else:
            fillcolor=None
        if fillcolor:
            self.Fillcolor=fillcolor
        if Text:
            self.Text=Text
        if AltText:
            self.AlternativeText=AltText
        self.updatecontrol = True
        self.Worksheet.drawShape(self)
        self.updatecontrol = False
        
    def Delete(self):
        self.Worksheet.deleteshape(self)
        pass
    
    def ZOrder(self,zorder):
        self.ZOrder_Val=zorder
    
    def set_activeflag(self,value):
        self.Active=value
        
    def get_activeflag(self):
        try:
            return self.Active
        except:
            self.Active=True
        return self.Active
    
    def Select(self):
        global Selection
        self.Worksheet.setSelectedShapes(self)
        Selection=CSelection(None,ws=self.Worksheet)
        
    def get_visible(self):
        return self.Visible_val
    
    def set_visible(self, value):
        self.Visible_val=value
        #self.Tshape.Visible = value
        if value==True:
            pass # redrawshape
        else:
            pass # removeshape

    Visible = property(get_visible, set_visible, doc='Visible Status')
    
    def get_name(self):
        return self.Name_val
    
    def set_name(self, value):
        self.Name_val=value
        #self.Tshape.Name=value

    Name = property(get_name, set_name, doc='Shape-Name')

    def get_left(self):
        return self.Left_val
    
    def set_left(self, value):
        self.Left_val=value
        #self.Tshape.Left=value

    Left = property(get_left, set_left, doc='Shape-Left')
    
    def get_top(self):
        return self.Top_val
    
    def set_top(self, value):
        self.Top_val=value

    Top = property(get_top, set_top, doc='Shape-Top')    
    
    def get_TopLeftCell(self):
        TopLeftCell = CRange(self.TopLeftCell_Row,self.TopLeftCell_Col)
        return TopLeftCell
    
    def set_TopLeftCell(self, value):
        pass
    
    TopLeftCell = property(get_TopLeftCell, set_TopLeftCell, doc='Shape-Name')

    def get_TopLeftCellRow(self):
        return self.TopLeftCell_Row_val
    
    def set_TopLeftCellRow(self, value):
        self.TopLeftCell_Row_val = value
    
    TopLeftCell_Row = property(get_TopLeftCellRow, set_TopLeftCellRow, doc='Shape-Name')

    def get_OnAction(self):
        return self.OnAction_val
    
    def set_OnAction(self, value):
        self.OnAction_val=value
        if self.control_dict:
            control_dict = self.control_dict.get("Components",None)
        else:
            control_dict = None
        if control_dict:
            button = control_dict[0]
            button_command = self.Worksheet.Workbook.jumptable.get(value, None) #D00.globalprocs.get(value,None)
            button["Command"]=button_command

    OnAction = property(get_OnAction, set_OnAction, doc='Shape-OnAction')
    
    def shape_button_1(self,event=None):
        #print("Shape Button_1 event")
        #print(repr(event))
        worksheet=self.Worksheet
        tags=worksheet.tksheet.MT.gettags(self.rectidx)
        #print(tags)
        if worksheet.left_click_callback:
            res_continue = worksheet.left_click_callback(tags, "",callertype="canvas")
        return
    
    def get_text(self):
        return self.Text_val
    
    def set_text(self, value):
        self.Text_val=value
        if self.rectidx!=0 or self.formwin!=None:
            worksheet=self.Worksheet # global_worksheet_dict[self.tablename]
            self.updatecontrol = True
            worksheet.drawShape(self, force=True) 
            self.updatecontrol = False
        
    Text = property(get_text, set_text, doc='Shape-Text')    

#*******************************************
# Excel Object Cells as DICT (old syntax Cells["A1:A5"])
#*******************************************        
class CCellDict(object):
    def __init__ (self,ws=None):
        #data = {}
        self.ws = ws
        
    def __getitem__(self, k):
        #print("Getitem",k)
        return Cells(k[0],k[1],ws=self.ws)
        
    def __setitem__(self,k,value):
        cell = Cells(k[0],k[1],ws=self.ws)
        cell.set_value(value)
        


#*******************************************
# Excel Object Range as DICT (old syntax Range["A1:A5"])
#*******************************************        
        
class CRangeDict(object):
    def __init__ (self,ws=None):
        self.ws = ws
        
    def __getitem__(self, k):
        #print("Getitem",k)
        if self.ws == None:
            self.ws= ActiveSheet
        return Range(k)
        
    def __setitem__(self,k,value):
        p_range=Range(k)
        p_range.Value=value
        
#*******************************************
# Excel Object Row as DICT (old syntax Row["A1:A5"])
#*******************************************        
        
class CRowDict(object):
    def __init__ (self,ws=None):
        #data = {}
        self.ws = ws
        
    def __getitem__(self, k):
        #print("Getitem",k)
        if self.ws == None:
            self.ws= ActiveSheet
        value = CRange(k)
        #value = self.ws.tksheet.get_data(k)
        return value # self.ws.Rows(k)
        
    def __setitem__(self,k,value):
        if self.ws == None:
            self.ws= ActiveSheet
        self.ws.tksheet.span(k).data = value
        #self.ws.wsRowdict[k] = value
        
        #print("CRowDict.Setitem",k, value)
        
#*******************************************
# Excel Object Sheets as DICT (old syntax Sheets["Main"])
#*******************************************        
        
class CSheetsDict(object):
    def __init__ (self,ws=None):
        #data = {}
        self.ws:CWorksheet = ws
        
    def __getitem__(self, k):
        #print("Getitem",k)
        if self.ws == None:
            self.ws= ActiveSheet
        value = self.ws.Workbook.Sheets(k)
        return value
        
    def __setitem__(self,k,value):
        pass

#*******************************************
# Excel Object WorksheetFunction https://learn.microsoft.com/en-us/office/vba/api/excel.worksheetfunction
#*******************************************        
        
class CWorksheetFunction(object):
    def __init__(self):
        pass
    
    def RoundUp(v1,v2):
        #print("RoundUp")
        if v2 == 0:
            if v1 == int(v1):
                return v1
            else:
                return int(v1 + 1)
        else:
            return "Error"
        
    def Power(self, p1,p2):
        return p1 ** p2
    
    def Max(self,*args):
        maxval=0
        for argvalue in args:
            value=val(argvalue)
            if value>maxval:
                maxval=value
        return maxval
    
    def Match(self,Lookup_value,Lookup_array, Match_type):
    # Returns the relative position of an item in an array that matches a specified value in a specified order. Use Match instead of one of the Lookup functions when you need the position of an item in a range instead of the item itself.    
    # Lookup_value: the value that you use to find the value that you want in a table.
    # Lookup_array: a contiguous range of cells containing possible lookup values. Lookup_array must be an array or an array reference.    
    # Match_type: the number -1, 0, or 1. Match_type specifies how Microsoft Excel matches lookup_value with values in lookup_array.
    # Match returns the position of the matched value within lookup_array, not the value itself. For example, MATCH("b",{"a","b","c"},0) returns 2, the relative position of "b" within the array {"a","b","c"}.
        res = Lookup_array.Find(Lookup_value)
        if res == None:
            return 0
        else:
            return res.Row
    
    def Min(self,*args):
        return min(*args)    

#*******************************************
# Excel Object Application https://learn.microsoft.com/en-us/office/vba/api/excel.application(object)
#*******************************************        
        
class CApplication(object):
    def __init__(self, parentframe=None, caption="", Path=None, p_global_controller=None):
        global ActiveApplication, ActiveCell, ActiveSheet, ActiveWindow,  ActiveWorkbook, Application, Selection, ThisCell, Worksheets,  Workbooks, global_controller 
        self.ActiveApplication = ActiveApplication = Application = self
        #self.ActiveCell = None # the Global ActiveCell is defined as a function
        self.ActiveSheet = None
        self.ActiveWindow = ActiveWindow = CActiveWindow()
        self.ActiveWorkbook = None
        Selection=CSelection()
        #*******************************************
        # Excel Object Application Properties
        #*******************************************        
        self.Calculation = xlCalculationAutomatic
        self.Caller=""
        self.Caption = caption
        self.controller = global_controller = p_global_controller
        self.EnableCalculation = True
        self.EnableEvents = True
        self.Height = 500
        self.hWnd = 0
        self.LanguageSettings = CLanguageSettings()
        self.Left = 0
        self.Name = caption
        self.Path = Path
        self.ScreenUpdating = True
        self.StatusBar = ""
        self.ThisCell = ThisCell = None
        self.ThisWorkbook = None
        self.Top = 0
        self.Version = "15"
        self.WindowState = 0
        self.Width = 1000
        self.WorksheetFunction = CWorksheetFunction()
        self.workbookslist = []
        self.Workbooks = Workbooks = CWorkbooks(application=self)
        self.worksheetslist = []
        self.Worksheets = Worksheets = CWorksheets(application=self)
        self.canvas_leftclickcmd=None
        self.parentframe = parentframe
        
        #*******************************************
        # Excel Object Application Methods
        #*******************************************                
    def GetOpenFilename(self,InitialFileName="", fileFilter="", Title="" ):
        filefilter_list = fileFilter.split(",")
        filepath = tk.filedialog.askopenfilename(filetypes=[filefilter_list],defaultextension=".MLL_pgf",title=Title,initialfile=InitialFileName)
        return filepath

    def GetSaveAsFilename(self,InitialFileName="", fileFilter="", Title="" ):
        filefilter_list = fileFilter.split(",")
        filepath = tk.filedialog.asksaveasfilename(filetypes=[filefilter_list],defaultextension=".MLL_pgf",title=Title,initialfile=InitialFileName)
        return filepath
    
    # Attention: Due to backwards compatibility: the Global ActiveCell is defined as a function
    def set_activecell(self, value):
        pass
    def get_activecell(self):
        global Selection
        row, col = self.ActiveSheet.getSelectedCell()
        activecell = self.ActiveSheet.Cells(row,col)
        self.Selection = Selection = CSelection(activecell,ws=self.ActiveSheet)
        return activecell 
    ActiveCell = property(get_activecell, set_activecell, doc = "ActiveCell")
    
    def set_activesheet(self, value):
        global ActiveSheet
        ActiveSheet = value
    def get_activesheet(self):
        return ActiveSheet
    ActiveSheet = property(get_activesheet, set_activesheet, doc = "ActiveSheet")
    
    def set_activeworkbook(self, value):
        global ActiveWorkbook
        ActiveWorkbook = value
    def get_activeworkbook(self):
        return ActiveWorkbook
    ActiveWorkbook = property(get_activeworkbook, set_activeworkbook, doc = "ActiveWorkbook")

    def Columns(self, column=None):
        pass

    def International(self,param):
        if param == xlDecimalSeparator:
            return ","
    
    def Intersect(self,rng1,rng2):
        if type(rng1)==CRange:
            rowrange1 = (rng1.start[0],rng1.end[0])
            colrange1 = (rng1.start[1],rng1.end[1])
        else:
            rowrange1 = rng1.rowrange
            colrange1 = rng1.colrange
            
        if type(rng2)==CRange:
            rowrange2 = (rng2.start[0],rng2.end[0])
            colrange2 = (rng2.start[1],rng2.end[1])
        else:
            rowrange2 = rng2.rowrange
            colrange2 = rng2.colrange
        
        #get rowrange intersection
        rowintersect=(max(rowrange1[0],rowrange2[0]),
                      min(rowrange1[1],rowrange2[1])+1)
                      
        colintersect=(max(colrange1[0],colrange2[0]),
                      min(colrange1[1],colrange2[1])+1)
        
        newrowrange=None
        if rowintersect[0]<rowintersect[1]:
            newrowrange=rowintersect
            
        newcolrange=None
        if colintersect[0]<colintersect[1]:
            newcolrange=colintersect
            
        if newcolrange !=None and newrowrange!=None:
            return CRange((newrowrange[0],newcolrange[0]),(newrowrange[1],newcolrange[1]))
        else:
            return None    
        
    def OnTime(self,time,cmd):
        #print("Application OnTime:",time,cmd)
        global_controller.after(time,cmd)
        return
    
    def Quit(self):
        pass
    
    def Range(self, Cell1=None, Cell2=None):
        pass
    
    def Rows(self, row=None):
        pass
    
    def Run(self,cmd):
        methodToCall = ActiveWorkbook.jumptable(cmd, None) #D00.globalprocs.get(cmd,None)
        if methodToCall:
            methodToCall()

    def RoundUp(self,v1,v2):
        #print("RoundUp")
        if v2 == 0:
            if v1 == int(v1):
                return v1
            else:
                return int(v1 + 1)
        else:
            return "Error"
    
    def setActiveWorkbook(self,workbookName):
        for wb in Workbooks:
            if wb.Name == workbookName:
                set_activeworkbook(wb)
                break
            
    def set_canvas_leftclickcmd(self,cmd):
        self.canvas_leftclickcmd=cmd
     
    def set_selection(self, value):
        pass
    def get_selection(self):
        pass
    Selection = property(get_selection, set_selection, doc = "Selection")

    
locale2msoid={"de":msoLanguageIDGerman,
              "en":msoLanguageIDEnglishUK,
              "nl":msoLanguageIDDutch,
              "fr":msoLanguageIDFrench,
              "it":msoLanguageIDItalian,
              "es":msoLanguageIDSpanish,
              "da":msoLanguageIDDanish
              }
msoid_list = ("de","en","nl","fr","it","es","da")

class CLanguageSettings(object):
    def __init__(self):
        id = -1
        
    def LanguageID(self,id):
        if id == msoLanguageIDUI:
            lngconfstr = global_controller.getConfigData("language")
            if lngconfstr == "":
                lngconfnum=0
            else:
                lngconfnum= int(lngconfstr)
            if lngconfnum==0:
                loc = locale.getdefaultlocale()
                locstr=loc[0]
                if locstr != None:
                    locstr=loc[0][:2]
                else:
                    locstr="de"
                msoid = locale2msoid.get(locstr,msoLanguageIDEnglishUK)
                return msoid
            else:
                msoid = locale2msoid.get(msoid_list[lngconfnum-1],msoLanguageIDEnglishUK)
                return msoid
        else:
            return msoLanguageIDEnglishUK
        
class CFont(object):
    def __init__(self,name="Font",size=8,shapetype=msoTextBox,shape=None, parent=None):
        self.Name = name
        self.parent = parent
        self.Size = size
        self.Shapetype= shapetype
        self._color_val = "#000000"
        #self.Color="#000000"
        self.Bold=False
        self.Italic=False
        self.Underline=- 4142
        self.Fill = CFill(self._color_val,shape=shape)
        
    @property
    def Color(self):
        return self._color_val
        
    @Color.setter
    def Color(self, value):
        self._color_val = value
        if self.parent == None:
            return
        else:
            if type(self.parent) == CRange:
                for cell in self.parent:
                    self.parent.ws.tksheet.highlight((cell.start[0]-1,cell.start[1]-1), fg=self._color_val)                
        
class CSelectedSheets(object):
    def _init__(self):
        pass
        
class CActiveWindow(object):
    def __init__(self):
        self.Zoom = 100
        self.ScrollColumn = 1
        self.SelectedSheets = {}
        self.SelectedSheets["Count"]=1
        
class SoundLines(object):
    def __init__(self):
        self.dict = {}
        self.Count = 0
        self.keys = []
    
    def Exists(self,Channel:int):
        pass
    
    def Add(Channel, Pin_playerClass):
        pass
    
class CCharacters(object):
    def __init__(self,text="",shape=None):
        self.Shape=shape
        #self.Text = text
        #self.Textrange.Text=text
        self.charformat_val = {} #dict(CCharFormat(0,len(text),shape)
        
    def get_text(self):
        return self.Shape.Text
    
    def set_text(self, value):
        self.Shape.Text=value
        
        #if self.Shape.Tshape:
        #    self.Shape.Tshape.Text=value

    Text = property(get_text, set_text, doc='Character-Text')
    
    def charformat(self,start,length,Bold=None,ForeColor=None):
        if length>0:
            #charformat = CCharFormat(start,length,self.Shape)
            
            charformat= CCharFormat(start, length, shape=self.Shape)
            if Bold!=None:
                if Bold==True:
                    charformat.Font="bold"
                else:
                    charformat.Font="normal"
            if ForeColor!=None:
                charformat.ForeGround=int2tkcolor(int(ForeColor))
            self.charformat_val[(start,length)] = charformat
            self.Shape.format_dict[self.Shape.Name] = self.charformat_val
            return

        
class CCharFormat(object):
    def __init__(self,start,length,shape=None):
        self.Font = None
        self.ForeGround = None
        

class CTextRange(object):
    def __init__(self,text="",shape=None):
        self.Shape=shape
        self.Characters = CCharacters(text=text,shape=shape)
        #self.Text = text
        self.Font = CFont(shape=shape)
        
    def get_text(self):
        #if self.Shape.Tshape:
        #    text=self.Shape.Tshape.Text
        #else:
        if hasattr(self.Shape,"Text"):
            text=self.Shape.Text
        else:
            text=""
        return text
    
    def set_text(self, value):
        
        self.Shape.Text=value
        
        #if self.Shape.Tshape:
        #    self.Shape.Tshape.Text=value

    Text = property(get_text, set_text, doc='Character-Text')             

class CTextFrame2(object):
    def __init__(self,text="",shape=None):
        self.TextRange = CTextRange(text=text,shape=shape)
    
class CButton(object):
    def __init__(self,name,shape=None):
        self.Shape = shape
        self.Name=name
        self.AlternativeText = ""
        self.TextFrame2 = CTextFrame2(shape=shape)
        self.Fill = (0,0,0)

        
class CControl(object):
    def __init__(self,value):
        self.Value=value

def Delay(dseconds):
    ActiveWorkbook.Delay(dseconds)
        
def rgb2tkcolor(r,g,b):
    return f'#{r:02x}{g:02x}{b:02x}'

def tkcolor2rgb(color):
    r = int(color[1:3], 16)
    g = int(color[3:5], 16)
    b = int(color[5:7], 16)
    return (r,g,b)

def tkcolor2int(color):
    r = int(color[1:3], 16)
    g = int(color[3:5], 16)
    b = int(color[5:7], 16)
    colorint=(b*256+g)*256+r
    return colorint

def int2tkcolor(colorVal):
    if type(colorVal)==int:
        if colorVal<0:
            _fn_return_value=""
        r=colorVal % 256
        g=(colorVal // 256 )  % 256
        b=colorVal // 65536
        _fn_return_value = f'#{r:02x}{g:02x}{b:02x}'
    else:
        _fn_return_value=colorVal
    return _fn_return_value

def get_global_controller():
    return ActiveWorkbook.controller

# global variables

#*******************************************
# Excel Statements https://learn.microsoft.com/en-us/office/vba/language/reference/statements
#*******************************************

def Unload(UserForm):
    UserForm.destroy()

def ChDrive(srcdir):
    if checkplatform("Windows"):
        drive = srcdir[0:2]
        os.chdir(drive)
    return

#*******************************************
# Excel Application properties https://learn.microsoft.com/en-us/office/vba/api/excel.application(object)
#*******************************************
Application = None

def ActiveCell():
    return ActiveApplication.ActiveCell
    global Selection
    row, col = ActiveSheet.getSelectedCell()
    activecell = ActiveSheet.Cells(row,col)
    Selection = CSelection(activecell,ws=ActiveSheet) #*HL
    return activecell #*HL

ActiveApplication = None


ActiveSheet: CWorksheet = None

ActiveWindow = CActiveWindow()

ActiveWorkbook: CWorkbook = None

Caption = ""

CellDict = CCellDict() # alias for Cells[] 

def Cells(row:int, column:int,ws=None):
    if ws == None:
        ws=ActiveSheet
    xlrange = ws.Cells(row,column)
    return xlrange

def Columns(col:int):
    ws = ActiveSheet
    lastusedrow=ws.LastUsedRow
    startcell=CRange(1,col)
    endcell  =CRange(lastusedrow,col)
    return CRange(startcell,endcell)

def Range(cell1=None,cell2=None,ws=None):
    if ws==None:
        ws=ActiveSheet
    if cell1==None:
        return CRange() # dummy for type definition only
    if type(cell1) == str:
        if ":" in cell1: # range as string a:b
            cellsplit = cell1.split(":")
            return Range(cellsplit[0],cellsplit[1])
        named_cell1 = ws.find_RangeName(cell1, ws=ws)
        if cell2 != None:
            if type(cell2) == str:
                named_cell2 = ws.find_RangeName(cell2, ws=ws)
                return CRange(named_cell1,named_cell2,ws=ActiveSheet)
            else:
                return CRange(named_cell1,cell2,ws=ActiveSheet)
        else:
            return CRange(named_cell1,named_cell1,ws=ActiveSheet)
    else:
        if cell2==None:
            cell2=cell1
        return CRange(cell1,cell2,ws=ActiveSheet)

def Rows(row=None,ws=None):
    if ws==None:
        ws=ActiveSheet
    if row==None:
        return CRange((1,1),(ws.getLastUsedRow(),1),ws=ws,parent=None,parenttype="Rows")
    else:
        if type(row)==str:
            rowlist=row.split(":")
            return CRange((int(rowlist[0]),1),(int(rowlist[1]),1),ws=ws,parent=None,parenttype="Rows")
        else:
            return CRange((row,1),(row,1),ws=ws,parent=None,parenttype="Rows")

RangeDict = CRangeDict() # alias for Range[]

RowDict = CRowDict() # alias for Rows[]


Selection = CSelection()
#Selection = property(get_selection, set_selection, doc='Selection') 


def Sheets(sheetname):
    return ActiveWorkbook.Sheets(sheetname)

Workbooks: CWorkbooks = []

Worksheets: CWorksheets = []

DefaultFont =  ("Calibri",10)

#*******************************************
# Excel Functions https://learn.microsoft.com/en-us/office/vba/language/reference/functions-visual-basic-for-applications
#*******************************************
def Date():
    x = datetime.datetime.now()
    return x.strftime("%x")

def Date_str():
    return Date()

def Dir(filepath,dummy=None):
    if os.path.isfile(filepath):
        return filepath
    else:
        if os.path.isdir(filepath):
            return filepath
        else:
            return ""
        
def DoEvents():
    #ActiveSheet.table.redraw()
    global_controller.update()

def Format(value,formatstring):
    if formatstring == "hh:mm:ss":
        time_val = time.gmtime(value)
        return time.strftime("%H:%M:%S",time_val)
    elif formatstring == "hh:mm":
        if type(value)==str:
            return value[:5]
        else:
            time_val = time.gmtime(value)
            return time.strftime("%H:%M",time_val)
    elif formatstring == "0000":
        return "{0:4d}".format(value)
    elif formatstring == "00":
        return "{0:2d}".format(int(value))
    elif formatstring == "dd.mm.yy":
        return value[:2]+"."+value[3:5]+"."+value[6:]
    else:
        pass
    return str(value) # no formating implemented yet

def GetAsyncKeyState(key):
    global shift_key
    if keyboard_Module_imported:
        if key==__VK_UP:
            return keyboard.is_pressed("up")
        if key==__VK_DOWN:
            return keyboard.is_pressed("down")
        if key==__VK_RETURN:
            return keyboard.is_pressed("enter")
        if key==__VK_ESCAPE:   
            return keyboard.is_pressed("escape")
        if key==__VK_CONTROL:   
            return keyboard.is_pressed("crtl")
        if key==__VK_SHIFT:
            Debug.Print("Check Shift Key")
            fn_return_value = shift_key #or keyboard.is_pressed("shift")
            shift_key=False
            return fn_return_value
        return False
    else:
        return False
"""
def InputBox(Message:str, Title:str, Default=None):
    #res = tk.simpledialog.askstring(Title,Message,initialvalue=Default,parent=PG.dialog_parent)
    F00.Userform_SimpleInput.Show(Message,Title,Default)
    res = F00.Userform_SimpleInput.UserForm_res
    if res == None:
        res=""
    return res
"""

def IsEmpty(obj):
    if len(obj)==0:
        return True
    else:
        return False
    
def IsError(testval):
    return False

def MsgBox(ErrorMessage:str, msg_type:int, ErrorTitle:str):
    if msg_type == vbQuestion + vbYesNoCancel or msg_type == vbYesNoCancel:
        res=tk.messagebox.askyesnocancel(title=ErrorTitle, message=ErrorMessage)
        if res == None:
            return vbCancel
        if res:
            return vbYes
        else:
            return vbNo
    elif msg_type == vbQuestion + vbYesNo or msg_type == vbYesNo:
        res=tk.messagebox.askyesno(title=ErrorTitle, message=ErrorMessage)
        if res == None:
            return vbCancel
        if res:
            return vbYes
        else:
            return vbNo        
    elif msg_type == vbOKCancel:
        res=tk.messagebox.askokcancel(title=ErrorTitle, message=ErrorMessage)
        if res == None:
            return vbCancel
        if res:
            return vbOK
        else:
            return vbCancel
    elif msg_type == vbOKOnly:
        res=tk.messagebox.showinfo(title=ErrorTitle, message=ErrorMessage)
        
    elif msg_type == vbInformation:
        res=tk.messagebox.showinfo(title=ErrorTitle, message=ErrorMessage)
        
    elif msg_type == vbCritical:
        res=tk.messagebox.showerror(title=ErrorTitle, message=ErrorMessage)
        
    elif msg_type == vbCritical + vbYesNo:
        res=tk.messagebox.askyesno(title=ErrorTitle, message=ErrorMessage)
        if res == None:
            return vbCancel
        if res:
            return vbYes
        else:
            return vbNo
    else:
        logging.debug("P01_MSGBox: Unknown Messagetype:"+str(msg_type))
        res=tk.messagebox.askyesnocancel(title=ErrorTitle, message=ErrorMessage)
        if res == None:
            return vbCancel
        if res:
            return vbYes
    return vbNo

def Run(cmd):
    subprocess.run(cmd,shell=True)    

def Shell(cmd):
    Run(cmd)
    
def Time():
    return time.time()

def TimeValue(Duration):
    return str(Duration)

def Time_str():
    t = datetime.datetime.now()
    return t.strftime("%X")

def VarType(obj):
    valtype=type(obj)
    if valtype is str:
        return vbString
    else:
        return None


    





